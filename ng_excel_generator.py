from __future__ import annotations

import argparse
from collections import defaultdict, Counter
from datetime import datetime, date
from pathlib import Path
import json
import os
import re
from typing import Any, cast, TYPE_CHECKING

from openpyxl import Workbook, load_workbook
from urllib.error import URLError, HTTPError
from urllib.parse import quote
from urllib.request import Request, urlopen


TRANSPORT_MAP = {
    "ocean": "Sea",
    "air": "Air",
    "sea": "Sea",
    "vessel": "Sea",
    # Arcteryx-specific transport codes
    "s1 - seafreight": "Sea",
    "s1": "Sea",
    "a1 - airfreight": "Air",
    "a1": "Air",
    "a2 - airfreight": "Air",
    "a2": "Air",
    "seafreight": "Sea",
    "airfreight": "Air",
    "sea freight": "Sea",
    "air freight": "Air",
    # Courier variants
    "courier": "Courier",
    "dhl": "Courier",
    "fedex": "Courier",
    "ups": "Courier",
    # M88 shorthand transport code
    "v": "Sea",
    "private parcel": "Courier",
    "private parcel service": "Courier",   # ON Running
    "parcel": "Courier",                   # Burton
    "international distributor": "Sea",    # Cotopaxi
    # Ocean carrier names (prAna and others use carrier name instead of mode)
    "maersk ocean": "Sea",
    "maersk": "Sea",
    "hapag-lloyd": "Sea",
    "hapag lloyd": "Sea",
    "msc": "Sea",
    "cma cgm": "Sea",
    "evergreen": "Sea",
    "cosco": "Sea",
    "yang ming": "Sea",
    "one": "Sea",       # Ocean Network Express
    # Hunter shipping instruction codes
    "sos - hunter sos": "Sea",
    "fb - hunter - fob warehouse": "Sea",
    "sms - sample warehouse": "Sea",
    "dte - davies turner e-com warehouse": "Sea",
    "hm - hammer gmbh & co. kg": "Sea",
    "hmcd - hammer cross dock": "Sea",
}

# Valid mapped transport values after TRANSPORT_MAP resolution
VALID_TRANSPORT_VALUES = {"Sea", "Air", "Courier"}

# Country code → full name (used for TransportLocation)
COUNTRY_NAME_MAP = {
    "AE": "UAE",
    "AR": "Argentina",
    "AT": "Austria",
    "AU": "Australia",
    "BR": "Brazil",
    "CA": "Canada",
    "CH": "Switzerland",       # ← BUG FIX (was "China")
    "CL": "Chile",
    "CN": "China",
    "CZ": "Czech Republic",
    "DE": "Germany",
    "DK": "Denmark",
    "EC": "Ecuador",
    "ES": "Spain",
    "FR": "France",
    "GB": "UK",
    "GR": "Greece",
    "HK": "Hong Kong",
    "HR": "Croatia",
    "HU": "Hungary",
    "ID": "Indonesia",
    "IL": "Israel",
    "IN": "India",
    "IT": "Italy",
    "JP": "Japan",
    "KR": "Korea",
    "MN": "Mongolia",
    "NP": "Nepal",
    "MT": "Malta",
    "MX": "Mexico",
    "MY": "Malaysia",
    "PA": "Panama",
    "PE": "Peru",
    "PH": "Philippines",
    "PL": "Poland",
    "RS": "Serbia",
    "RU": "Russia",
    "TH": "Thailand",
    "TR": "Turkey",
    "TW": "Taiwan",
    "UK": "UK",
    "US": "USA",
    "US WHOLESALE 3PL": "USA",
    "US RETAIL 3PL": "USA",
    "US ECOMM": "USA",
    "UNITED KINGDOM": "UK",
    "UNITED ARAB EMIRATES": "UAE",
    "UNITED STATES": "USA",
    "UY": "Uruguay",
    "VN": "Vietnam",
    "ZA": "South Africa",
    # Full country name passthrough normalization (e.g. Haglofs destination column)
    "SWEDEN": "Sweden",
    "KOREA": "Korea",
    "JAPAN": "Japan",
    "HONG KONG": "Hong Kong",
    "GERMANY": "Germany",
    "FRANCE": "France",
    "ITALY": "Italy",
    "SPAIN": "Spain",
    "NETHERLANDS": "Netherlands",
    "BELGIUM": "Belgium",
    "SWITZERLAND": "Switzerland",
    "AUSTRIA": "Austria",
    "DENMARK": "Denmark",
    "NORWAY": "Norway",
    "FINLAND": "Finland",
    "POLAND": "Poland",
    "CZECH REPUBLIC": "Czech Republic",
    "AUSTRALIA": "Australia",
    "CANADA": "Canada",
    "CHINA": "China",
    "INDIA": "India",
    "INDONESIA": "Indonesia",
    "MALAYSIA": "Malaysia",
    "THAILAND": "Thailand",
    "VIETNAM": "Vietnam",
    "TAIWAN": "Taiwan",
    "SINGAPORE": "Singapore",
    "CZECHIA": "Czech Republic",          # Burton
    "GREAT BRITAIN": "UK",                 # Hunter
    "TBC": "",                             # Hunter — unknown destination, leave blank
}

CURRENCY = "USD"
SUPPLIER_PROFILE = "DEFAULT_PROFILE"
DEFAULT_CUSTOMER = "COL"

# Plant name / plant code → TransportLocation country
# Used for M88 buy files that have no explicit TransportLocation column.
PLANT_COUNTRY_MAP: dict[str, str] = {
    # Plant name patterns (lowercase)
    "visalia dc":             "USA",
    "visalia":                "USA",
    "jonestown dc":           "USA",
    "jonestown":              "USA",
    "brampton dc":            "Canada",
    "brampton":               "Canada",
    "dropship us":            "USA",
    "dropship international": "USA",
    "dropship dc":            "USA",
    "dropship ca":            "Canada",
    "vf outdoor mexico":      "Mexico",
    "vf outdoor mexico s de r l d": "Mexico",
    "photoshooting":          "BELGIUM",
    "eu main":                "BELGIUM",
    "eu uk":                  "UK",
    "eu":                     "EU",
    "japan":                  "Japan",
    "korea":                  "Korea",
    "australia":              "Australia",
    "hong kong":              "Hong Kong",
    "china":                  "China",
    "virtual":                "Dubai",
    "argentina":              "Argentina",
    "brazil":                 "Brazil",
    "chile":                  "Chile",
    "guatemala":              "Guatemala",
    "panama":                 "Panama",
    "peru":                   "PERU",
    "uruguay":                "URUGUAY",
    "united arab emirates":   "UNITED ARAB EMIRATES",
    "singapore":              "Singapore",
    "apdindc":                "Singapore",
    "israel":                 "Israel",
    "south africa":           "South Africa",
    "taiwan":                 "Taiwan",
    "thailand":               "Thailand",
    "malaysia":               "Malaysia",
    "nepal":                  "Nepal",
    "indonesia":              "Indonesia",
    # Plant code patterns
    "1001": "USA",
    "1010": "USA",
    "1020": "USA",
    "1004": "Canada",
    "1009": "USA",
    "1005": "Mexico",
    "t909": "Japan",
    "d060": "BELGIUM",
    "d080": "UK",
    "vd60": "Dubai",
    # Fox Racing plant codes — file uses no leading zeros (10, 11, 40, 50, 60)
    "0010": "",         # TODO: confirm destination
    "0011": "",         # TODO: confirm destination
    "0040": "",         # TODO: confirm destination
    "0050": "",         # TODO: confirm destination
    "0060": "",         # TODO: confirm destination
    "10":   "",         # Fox Racing — destination TBD
    "11":   "",         # Fox Racing — destination TBD
    "40":   "",         # Fox Racing — destination TBD
    "50":   "",         # Fox Racing — destination TBD
    "60":   "",         # Fox Racing — destination TBD
    # 511 Tactical WH codes
    "3020": "Sweden",   # SE.PO- prefix confirms Sweden
    "5001": "Hong Kong",  # HK.PO- prefix confirms Hong Kong
    # Vans DC Plant codes
    "1023": "USA",
    "d010": "Czech Republic",
    "vd10": "UAE",
    "d00028": "UAE",
    # Vans DC Plant name patterns
    "south ontario dc": "USA",
    "canada brampton dc": "Canada",
    "vf prague dc cz": "Czech Republic",
    "vf northern europe": "UK",
    "vf northern europe(uk)": "UK",
    "sun and sand sports": "UAE",
    "sun and sand sports llc": "UAE",
}

# ─────────────────────────────────────────────────────────────────────────────
# Brand-level supplier fallback map.
# When buy file does not have an explicit supplier/vendor column or when
# vendor_name is provided instead of a code, we resolve it here.
# Keys are lowercase brand identifiers found in the buy file.
# ─────────────────────────────────────────────────────────────────────────────
BRAND_SUPPLIER_MAP: dict[str, str] = {
    "col": "MSO",
    "columbia": "MSO",
    "tnf": "PT. UWU JUMP INDONESIA",
    "the north face": "PT. UWU JUMP INDONESIA",
    "arcteryx": "PT. UWU JUMP INDONESIA",
    "arc'teryx": "PT. UWU JUMP INDONESIA",
    "fox racing": "PT. UWU JUMP INDONESIA",
    "511 tactical": "PT. UWU JUMP INDONESIA",
    "haglofs": "PT. UWU JUMP INDONESIA",
    "mammut": "PT. UWU JUMP INDONESIA",
    "obermeyer": "Hangzhou U-Jump Arts and Crafts",
    "on running": "PT. UWU JUMP INDONESIA",
    "on ag": "PT. UWU JUMP INDONESIA",
    "66 degrees north": "PT. UWU JUMP INDONESIA",
    "peak performance": "PT. UWU JUMP INDONESIA",
    "prana": "PT. UWU JUMP INDONESIA",
    "burton": "PT. UWU JUMP INDONESIA",
    "cotopaxi": "PT. UWU JUMP INDONESIA",
    "hunter": "PT. UWU JUMP INDONESIA",
    "rossignol": "PT. UWU JUMP INDONESIA",
}

# Brand → friendly customer name used in output files.
BRAND_CUSTOMER_MAP: dict[str, str] = {
    "col": "Columbia",
    "columbia": "Columbia",
    "tnf": "The North Face In-Line",
    "the north face": "The North Face In-Line",
    "arcteryx": "Arcteryx",
    "arc'teryx": "Arcteryx",
    "haglofs": "Haglofs",
    "obermeyer": "Obermeyer",
    "on running": "On AG",
    "on ag": "On AG",
    "66 degrees north": "66 Degrees North",
    "peak performance": "Peak Performance",
    "prana": "prAna",
    "burton": "Burton",
    "cotopaxi": "Cotopaxi",
    "fox racing": "Fox Racing",
    "mammut": "Mammut",
    "rossignol": "Rossignol",
    "vans": "Vans",
    "south ontario dc": "Vans",
    "canada brampton dc": "Vans",
    "vf prague dc cz": "Vans",
    "vf northern europe": "Vans",
    "vf northern europe(uk)": "Vans",
    "sun and sand sports": "Vans",
    "sun and sand sports llc": "Vans",
    # Hunter: customer name in file is "Batra Group" — pass through raw, no override
    # "hunter": "Hunter",
    # Fox Racing: multiple distributors in file — customer passes through raw from file, no override
    # 511 Tactical: no customer column — use manual_customer at upload time
}

# ─────────────────────────────────────────────────────────────────────────────
# Customer subtype overrides.
# Keys are lowercase brand + subtype signal found in source data.
# Used by _resolve_customer_subtype() to map raw customer values that contain
# "RTO", "SMU", or other signals to the correct spelled-out customer name.
# ─────────────────────────────────────────────────────────────────────────────
TNF_CUSTOMER_SUBTYPE_MAP: dict[str, str] = {
    "the north face in-line": "The North Face In-Line",
    "the north face inline":  "The North Face In-Line",
    "the north face rto":     "The North Face RTO",
    "the north face smu":     "The North Face SMU",
    "tnf in-line":            "The North Face In-Line",
    "tnf inline":             "The North Face In-Line",
    "tnf rto":                "The North Face RTO",
    "tnf smu":                "The North Face SMU",
}

# ─────────────────────────────────────────────────────────────────────────────
# KeyUser map — per brand, the MLO team members that go into ORDERS.
# Keys are lowercase brand identifiers.
# Values are dicts keyed by KeyUser column name.
# Only KeyUser1 (Planning), KeyUser2 (Purchasing), KeyUser4 (Production),
# KeyUser5 (Logistics/Shipping) are populated per BRD; 3/6/7/8 left blank.
# ─────────────────────────────────────────────────────────────────────────────
BRAND_KEYUSER_MAP: dict[str, dict[str, str]] = {
    "tnf": {
        "KeyUser1": "Ron",
        "KeyUser2": "Maricar",
        "KeyUser3": "",
        "KeyUser4": "Ron",
        "KeyUser5": "Elaine Sanchez",
        "KeyUser6": "",
        "KeyUser7": "",
        "KeyUser8": "",
    },
    "the north face": {
        "KeyUser1": "Ron",
        "KeyUser2": "Maricar",
        "KeyUser3": "",
        "KeyUser4": "Ron",
        "KeyUser5": "Elaine Sanchez",
        "KeyUser6": "",
        "KeyUser7": "",
        "KeyUser8": "",
    },
    "col": {
        "KeyUser1": "",
        "KeyUser2": "",
        "KeyUser3": "",
        "KeyUser4": "",
        "KeyUser5": "",
        "KeyUser6": "",
        "KeyUser7": "",
        "KeyUser8": "",
    },
    "columbia": {
        "KeyUser1": "",
        "KeyUser2": "",
        "KeyUser3": "",
        "KeyUser4": "",
        "KeyUser5": "",
        "KeyUser6": "",
        "KeyUser7": "",
        "KeyUser8": "",
    },
    "arcteryx": {
        "KeyUser1": "",
        "KeyUser2": "",
        "KeyUser3": "",
        "KeyUser4": "",
        "KeyUser5": "",
        "KeyUser6": "",
        "KeyUser7": "",
        "KeyUser8": "",
    },
    "arc'teryx": {
        "KeyUser1": "",
        "KeyUser2": "",
        "KeyUser3": "",
        "KeyUser4": "",
        "KeyUser5": "",
        "KeyUser6": "",
        "KeyUser7": "",
        "KeyUser8": "",
    },
    "rossignol": {
        "KeyUser1": "Via",
        "KeyUser2": "April Joy",
        "KeyUser3": "",
        "KeyUser4": "Via",
        "KeyUser5": "Elaine Sanchez",
        "KeyUser6": "",
        "KeyUser7": "",
        "KeyUser8": "",
    },
    "vuori": {
        "KeyUser1": "Patrick",
        "KeyUser2": "Mary",
        "KeyUser3": "",
        "KeyUser4": "Patrick",
        "KeyUser5": "Elaine Sanchez",
        "KeyUser6": "",
        "KeyUser7": "",
        "KeyUser8": "",
    },
}

# Default KeyUser block (all blank) used when brand not found in map.
_DEFAULT_KEYUSERS: dict[str, str] = {
    k: "" for k in [
        "KeyUser1", "KeyUser2", "KeyUser3", "KeyUser4",
        "KeyUser5", "KeyUser6", "KeyUser7", "KeyUser8",
    ]
}

# ─────────────────────────────────────────────────────────────────────────────
# Template map — per brand, the NG template name to use.
# Overrides _normalize_template() default of "BULK" for brands with
# brand-specific template names (e.g. TNF uses "Major Brand Bulk").
# For LINES the template field uses a different, more granular value.
# ─────────────────────────────────────────────────────────────────────────────
BRAND_ORDERS_TEMPLATE_MAP: dict[str, str] = {
    "tnf":            "Major Brand Bulk",
    "the north face": "Major Brand Bulk",
    "col":            "BULK",
    "columbia":       "BULK",
    "arcteryx":       "BULK",
    "arc'teryx":      "BULK",
    "rossignol":      "Major Brand Bulk",
    "vuori":          "Major Brand Bulk",
}

BRAND_LINES_TEMPLATE_MAP: dict[str, str] = {
    "tnf":            "FOB Bulk EDI PO (New)",
    "the north face": "FOB Bulk EDI PO (New)",
    "col":            "BULK",
    "columbia":       "BULK",
    "arcteryx":       "BULK",
    "arc'teryx":      "BULK",
    "rossignol":      "FOB Bulk EDI PO (New)",
    "vuori":          "FOB Bulk Non EDI PO (New)",
}

# ─────────────────────────────────────────────────────────────────────────────
# Remote brand config (TS API) — optional source of truth.
# Python CLI will call this when available, otherwise fall back to maps above.
# ─────────────────────────────────────────────────────────────────────────────
_BRAND_CONFIG_CACHE: dict[str, dict[str, Any] | None] = {}


def _fetch_brand_config(brand: str) -> dict[str, Any] | None:
    base_url = os.getenv("BRAND_CONFIG_URL", "http://localhost:3000/api/brand-config")
    if not base_url or not brand:
        return None
    url = f"{base_url}?brand={quote(brand)}"
    try:
        req = Request(url, headers={"Accept": "application/json"})
        with urlopen(req, timeout=2) as resp:
            if resp.status != 200:
                return None
            payload = json.loads(resp.read().decode("utf-8"))
            return payload if isinstance(payload, dict) else None
    except (HTTPError, URLError, ValueError, OSError):
        return None


def _get_brand_config(brand: str) -> dict[str, Any] | None:
    key = (brand or "").strip().lower()
    if not key:
        return None
    if key in _BRAND_CONFIG_CACHE:
        return _BRAND_CONFIG_CACHE[key]
    config = _fetch_brand_config(brand)
    _BRAND_CONFIG_CACHE[key] = config
    return config

# Fallback fixed positions for the original COL buy file (1-based Excel columns)
DEFAULT_COL_MAP = {
    "po": 7,
    "product": 21,
    "product_alt": 16,
    "size": 0,
    "colour": 24,
    "qty": 28,
    "orig_ex_fac": 34,
    "trans_cond": 31,
    "buy_date": 9,
    "customer": 2,
    "brand": 2,
    "season": 6,
    "template": 8,
}

# ─────────────────────────────────────────────────────────────────────────────
# HEADER_ALIASES  –  extended to cover Arcteryx/Madison88, COL/INFOR, TNF
# ─────────────────────────────────────────────────────────────────────────────
HEADER_ALIASES: dict[str, list[str]] = {
        "transport_location": [
            "transportlocation", "transport location",
            "destination", "dest country", "ult. destination",
            # ON Running / Burton
            "country/region",
            # Cotopaxi raw INFOR export (2-letter country code column)
            "country",
            # Req 1: expanded transportLocation aliases
            "ship to country",
        ],
    "plant": [
        "plant", "plant code", "dc plant",
        # 511 Tactical
        "wh",
        # Vans
        "jde plant", "jde_plant", "dc plant",
    ],
    "plant_name": [
        # Madison88/TNF buy file
        "plant name", "ship-to party name",
        # Vans
        "destination name", "destination", "plant_name",
    ],
    "po": [
        "po #", "po#", "po", "pono", "po_number", "ponumber", "buyer po", "buyer po number",
        "purchase order", "purchaseorder",
        "extraction po #", "extraction po#",
        # Arcteryx – per-shipment tracking code used as PO key
        "tracking number",
        "master po#", "master po #",
        # Fox Racing
        "purchasing document number", "purchasing document",
        # 66 Degrees North
        "purchase order number",
        # Peak Performance
        "buy 1 - tracking no.",
        # Haglofs / Hunter / Book2
        "po number",
    ],
    "buyer_po_number": [
        "master po#", "master po #",
        # Req 1: expanded buyerPoNumber aliases
        "buyer po", "buyer po #", "customer po", "po#",
    ],
    "product": [
        "material style", "product", "style number", "style no", "style no.",
        "product name",
        # Arcteryx – Article is the colourway-level style code
        "article",
        # Generic fallbacks — "sku" and "item" removed: too generic, cause false matches
        "style",
        # Madison88/TNF buy file
        "material",
        # 511 Tactical
        "item#", "item #",
        # ON Running
        "buyer item #", "buyer item#",
        # Peak Performance
        "article code [sap]",
        # prAna
        "style #",
        # Hunter
        "item code",
    ],
    "product_alt": [
        "jde style",
        "m88 ref",
        # Arcteryx model = base style without colour suffix
        "model",
        # Madison88/TNF buy file base style
        "style#",
        # Peak Performance
        "model code [sap]",
        # ON Running
        "style",
    ],
    "product_external_ref": [
        "name",
        "product external ref",
        # 511 Tactical — human-readable description, not the style code
        "style name",
        # ON Running / 66 Degrees North / prAna / Cotopaxi / Hunter
        "short description", "style description", "description",
        "item description",
    ],
    "product_customer_ref": [
        "buyer style number",
        "buyer style no",
        "buyer style #",
        "buyer style",
        "product customer ref",
    ],
    "product_name": [
        "model description", "article name", "sku description",
        "material name", "description",
        "buyer style name",
    ],
    "size": [
        "size", "size name", "sizename", "product size", "productsize",
        "size code", "size #", "size#", "size_name", "size-name",
        # Madison88/TNF buy file
        "dim 1", "dim1",
        # Fox Racing
        "grid value",
        # Cotopaxi
        "merch - size",
    ],
    "customer": [
        "customer", "customer name", "brand",
        # Arcteryx
        "business unit description",
        # PLM/M88
        "plm customer name",
    ],
    "brand": [
        "brand",
        # Arcteryx
        "business unit description",
    ],
    "vendor_name": [
        "vendor name", "vendorname", "supplier name",
        "factory",
        "final vendor name",  # Madison88/TNF buy file
        # Fox Racing
        "goods supplier name",
        # Peak Performance
        "production supplier name",
    ],
    "vendor_code": [
        "vendor code", "vendorcode", "vendor", "supplier",
        "product supplier", "productsupplier",
        # 511 Tactical
        "updated fty",
        # Madison88/TNF buy file
        "final vendor", "final factory",
        # Req 1: expanded productSupplier aliases
        "supplier code", "mfr code",
    ],
    "season": [
        "season", "range", "productrange",
        # Madison88/TNF buy file
        "season indicator",
        # Hunter — Requisition No contains season+destination code e.g. "AW26_UKSOS"
        "requisition no",
        # Cotopaxi
        "merch - season",
        # Req 1: expanded season aliases
        "season code",
    ],
    "template": [
        "doc type", "template",
    ],
    "colour": [
        "color", "colour",
        # Arcteryx – Article Name holds the colour description
        "article name", "color description",
        # Madison88/TNF buy file colourway code lives in Material (used for PLM lookup key)
        "material",
        # Fox Racing — material description is the colour name; material itself maps to product
        "material description",
        # 511 Tactical
        "color code",
        # Haglofs — combined colour code+name (e.g. "5RA Bright Red")
        "style color",
        # Peak Performance
        "primary color peak pdm code",
        # Cotopaxi
        "merch - color",
        # Hunter
        "colour code",
        # Req 1: expanded colour aliases
        "colour desc", "colour description", "color desc", "color description",
    ],
    "colour_display": [
        # Madison88/TNF buy file human-readable colour name (output only, not PLM lookup key)
        "longtext",
        # prAna / Hunter
        "color", "colour description",
    ],
    "qty": [
        "ordered qty", "quantity", "qty",
        "sum of order total qty", "sum of qty (lum)", "sum of qty (lum)2",
        "open qty (pcs/prs)",
        # Arcteryx
        "requested qty",
        # Madison88/TNF buy file
        "final qty", "revised qty",
        # Fox Racing
        "order qty",
        "sum of order qty",
        # Peak Performance
        "final po qty",
        # EVO
        "bulk qty",
        # Req 1: expanded quantity aliases
        "qty ordered", "total qty", "units",
    ],
    "orig_ex_fac": [
        "orig ex fac", "delivery date", "deliverydate",
        "negotiated ex fac date", "ex fac",
        # Arcteryx
        "ex-factory",
        # Madison88/TNF buy file — Vendor Confirmed CRD is primary date source
        "vendor confirmed crd",
        "final crd (order date + lt1)", "brand requested crd",
        # Fox Racing
        "ex factory date",
        # 511 Tactical
        "updated planned exit date",
        # Peak Performance
        "buy 1 cfm crd",
        # Cotopaxi
        "requested exw date",
        # Burton
        "ex-factory date", "ex factory",
        # Hunter
        "efd",
        # ON AG (INFOR export)
        "ship window end date",
        # Req 1: expanded exFtyDate aliases
        "ship date", "ship window", "planned ship date",
        "in-dc date", "in dc date", "dc arrival date", "shipping date",
    ],
    "brand_requested_crd": [
        "brand requested crd",
    ],
    "vans_confirmed_vendor_crd": [
        "confirmed crd dt (vendor) -(vendor confirmed crd dt)",
    ],
    "confirmed_ex_fac": [
        "confirmed fty ex fac", "confirmed ex fac", "fty ex fac",
        # 511 Tactical
        "confirmed x-fty",
        # 66 Degrees North
        "delivery date",
    ],
    "trans_cond": [
        "trans cond", "transport method", "transportmethod",
        # Arcteryx
        "transport mode",
        # Madison88/TNF buy file
        "shipment mode", "transportation mode description",
        "transportation mode",
        # Fox Racing / Hunter
        "shipping instructions",
        # Burton
        "ship mode description", "ship mode",
        # prAna
        "ship via",
        # ON AG (INFOR export)
        "shipment method",
        # Req 1: expanded transportMethod aliases
        "mode", "freight mode", "shipping method",
    ],
    "buy_date": [
        "buy date", "keydate", "po issuance date",
        # Arcteryx
        "file date",
        # Madison88/TNF buy file
        "order date",
        # Fox Racing
        "purchasing document date",
        # 66 Degrees North / prAna
        "po date",
        # ON Running
        "create date",
        # Req 1: expanded poIssuanceDate aliases
        "issue date",
    ],
    "cancel_date": [
        "cancel date", "canceldate", "cancel", "udf-canel_date",
        # Fox Racing
        "so order cancel date",
        # Req 1: expanded cancelDate aliases
        "cancellation date",
    ],
    "status": [
        "status", "confirmation status",
        # Arcteryx
        "gsc type",
        # Fox Racing
        "po status",
    ],
    "submit_buy": [
        "submit buy", "buy round",
    ],
    "category": [
        "product group description", "product line description",
        "planning category", "dept", "department",
        "capacity type",
        # Req 1: expanded category aliases
        "division", "product division", "gender", "gender code",
    ],
    # ── Haglofs additions ──
    # (delivery date already covered by orig_ex_fac aliases above)
    # ── Fox Racing additions ──
    # purchasing document number → po (added to "po" key below via patch)
    # ── 511 Tactical additions ──
    # (item#, color code, wh, updated fty handled via po/product/colour/plant/vendor_name keys)
}

ORDERS_HEADERS = [
    "PurchaseOrder", "ProductSupplier", "Status", "Customer",
    "TransportMethod", "TransportLocation", "PaymentTerm", "Template",
    "KeyDate", "ClosedDate", "DefaultDeliveryDate", "Comments", "Currency",
    "KeyUser1", "KeyUser2", "KeyUser3", "KeyUser4", "KeyUser5",
    "KeyUser6", "KeyUser7", "KeyUser8",
    "ArchiveDate", "PurchaseUOM", "SellingUOM",
    "ProductSupplierExt", "FindField_ProductSupplier",
]

LINES_HEADERS = [
    "PurchaseOrder", "LineItem", "ProductRange", "Product", "Customer",
    "DeliveryDate", "TransportMethod", "TransportLocation", "Status",
    "PurchasePrice", "SellingPrice", "Template", "KeyDate", "SupplierProfile",
    "ClosedDate", "Comments", "Currency", "ArchiveDate",
    "ProductExternalRef", "ProductCustomerRef", "PurchaseUOM", "SellingUOM",
    "UDF-buyer_po_number", "UDF-start_date", "UDF-canel_date",
    "UDF-Inspection result", "UDF-Report Type", "UDF-Inspector",
    "UDF-Approval Status", "UDF-Submitted inspection date",
    "FindField_Product",
]

SIZES_HEADERS = [
    "PurchaseOrder", "LineItem", "Range", "Product",
    "SizeName", "ProductSize", "Quantity", "Colour", "Customer",
    "Department", "CustomAttribute1", "CustomAttribute2", "CustomAttribute3",
    "LineRatio", "ColourExt", "CustomerExt", "DepartmentExt",
    "CustomAttribute1Ext", "CustomAttribute2Ext", "CustomAttribute3Ext",
    "ProductExternalRef", "ProductCustomerRef",
    "FindField_Colour", "FindField_Customer", "FindField_Department",
    "FindField_CustomAttribute1", "FindField_CustomAttribute2",
    "FindField_CustomAttribute3", "FindField_Product",
]

# Expected column counts for output validation
EXPECTED_COL_COUNTS = {
    "ORDERS":      len(ORDERS_HEADERS),   # 26
    "LINES":       len(LINES_HEADERS),    # 31
    "ORDER_SIZES": len(SIZES_HEADERS),    # 29
}


# ─────────────────────────────────────────────────────────────────────────────
# Utility helpers
# ─────────────────────────────────────────────────────────────────────────────

def _as_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


PRODUCT_SHEET_ALIASES = {
    "color name": "colour",
    "colour name": "colour",
    "color": "colour",
    "colour": "colour",
    "factory": "factory",
    "vendor code": "factory",
    "vendorcode": "factory",
    "cost": "cost",
    "customer name": "customer_name",
    "customer": "customer_name",
    "product name": "product_name",
    "product": "product_name",
    "buyer style number": "buyer_style_number",
    "buyer style no": "buyer_style_number",
    "buyer style #": "buyer_style_number",
    "buyer style": "buyer_style_number",
}


def _normalize_colour_key(value: Any) -> str:
    raw = _as_text(value).strip()
    if not raw:
        return ""
    m = re.search(r"^[A-Z]{2,5}\s*-\s*([A-Z0-9]{2,5})\b", raw.upper())
    if m:
        return m.group(1)
    upper = raw.upper()
    compact = re.sub(r"\s*-\s*", "-", upper)
    compact = re.sub(r"\s+", "", compact)

    # TNF PLM colour names are typically "TNF-<CODE>-<DESC>".
    m = re.search(r"(?:^|-)TNF-([A-Z0-9]{2,4})(?:-|$)", compact)
    if m:
        return m.group(1)

    # Material / style variants often carry the colourway code as the trailing suffix.
    # e.g. "NF0A887VVQ2" -> "VQ2", "NF0A8HMENLQ" -> "NLQ".
    m = re.search(r"[A-Z0-9]{3,}([A-Z0-9]{3})$", compact)
    if m:
        return m.group(1)

    raw_l = raw.lower().strip()
    if re.fullmatch(r"\d+(\.\d+)?", raw_l):
        try:
            return str(int(float(raw_l)))
        except ValueError:
            return raw_l
    m = re.search(r"\d+", raw_l)
    if m:
        return m.group(0).lstrip("0") or "0"
    return raw_l


def _extract_style_colour_code(value: Any) -> str:
    upper = _as_text(value).strip().upper()
    m = re.search(r"([A-Z0-9]{3})$", upper)
    return m.group(1) if m else ""


def _normalize_style_key(value: Any) -> str:
    raw = _as_text(value).strip()
    if not raw:
        return ""
    upper = raw.upper()
    if isinstance(upper, str) and upper.startswith("NF0A") and len(upper) >= 5:
        return upper[-5:]  # type: ignore
    return raw


def _detect_product_sheet(ws: Any) -> tuple[bool, int]:
    header_row = 1
    best_score = -1
    buy_headers = {
        "po #", "pono", "purchase order", "purchaseorder", "line #s",
        "lineitem", "ordered qty", "qty", "quantity", "season", "brand",
        "dc plant", "ship-to party name",
        "sum of order total qty", "sum of qty (lum)2",
    }
    for r in range(1, min(50, ws.max_row) + 1):
        row_cells = [cell for cell in ws[r]]
        row_vals = [_as_text(c.value).lower().strip() for c in row_cells]
        
        row_product_score = len([v for v in row_vals if v in PRODUCT_SHEET_ALIASES])
        row_buy_score = len([v for v in row_vals if v in buy_headers])
        
        if row_buy_score >= 2:
            continue
        if row_product_score > best_score:
            best_score = row_product_score
            header_row = r
    # Exclude sheets that look like buy sheets (high buy_score means it's a buy file sheet)
    if best_score < 3:
        return False, header_row
    # Re-check buy_score at the detected header row
    header_row_tuple = cast(tuple[Any, ...], ws[header_row])
    header_vals = [_as_text(cell.value).lower().strip() for cell in header_row_tuple]
    final_header_buy_score = len([v for v in header_vals if v in buy_headers])
    
    if final_header_buy_score >= 2:
        return False, header_row
    return True, header_row


def _extract_product_sheet_map_from_wb(wb: Any) -> dict[str, list[dict[str, Any]]]:
    result: defaultdict[str, list[dict[str, Any]]] = defaultdict(list)
    seen_entries: set[tuple] = set()  # deduplicate across worksheets
    for ws in wb.worksheets:
        is_product, header_row = _detect_product_sheet(ws)
        if not is_product:
            continue
        header_map: dict[str, int] = {}
        header_row_tuple = cast(tuple[Any, ...], ws[header_row])
        for cell in header_row_tuple:
            if cell is None: continue
            val = cell.value
            if val is None: continue
            key = _as_text(val).lower().strip()
            mapped = PRODUCT_SHEET_ALIASES.get(key)
            if mapped and mapped not in header_map:
                col_val = cell.column
                if col_val is not None:
                    header_map[mapped] = col_val
        if "colour" not in header_map:
            continue
        for r in range(header_row + 1, ws.max_row + 1):
            colour_raw = _as_text(ws.cell(row=r, column=header_map["colour"]).value)
            colour_key = _normalize_colour_key(colour_raw)
            buyer_style_number = _as_text(ws.cell(row=r, column=header_map.get("buyer_style_number", 0)).value) \
                if header_map.get("buyer_style_number") else ""
            if not colour_key or not buyer_style_number:
                continue
            normalized_style = _normalize_style_key(buyer_style_number)
            entry = {
                "colour": colour_raw,
                "factory": _as_text(ws.cell(row=r, column=header_map.get("factory", 0)).value)
                if header_map.get("factory") else "",
                "cost": ws.cell(row=r, column=header_map.get("cost", 0)).value
                if header_map.get("cost") else "",
                "customer_name": _as_text(ws.cell(row=r, column=header_map.get("customer_name", 0)).value)
                if header_map.get("customer_name") else "",
                "product_name": _as_text(ws.cell(row=r, column=header_map.get("product_name", 0)).value)
                if header_map.get("product_name") else "",
                "buyer_style_number": buyer_style_number,
            }
            # Build lookup keys: use raw buyer_style_number AND each slash-separated segment
            # e.g. "217554/CU2279" → also index under "CU2279" to match buy file JDE Style
            # Exact matches are inserted first so they win over slash-segment duplicates
            lookup_keys_ordered = [(buyer_style_number, True)]  # (key, is_exact)
            if normalized_style and normalized_style != buyer_style_number:
                lookup_keys_ordered.append((normalized_style, False))
            for part in re.split(r"\s*/\s*", buyer_style_number):
                part = part.strip()
                if part and part != buyer_style_number:
                    lookup_keys_ordered.append((part, False))
            # Strip " - SUFFIX" or " (SUFFIX)" → "A8HME - LP5L"→"A8HME", "A8CGZ (A3FJW)"→"A8CGZ"
            style_base = re.split(r'\s*[\(\-]', buyer_style_number)[0].strip()
            if style_base and style_base != buyer_style_number:
                lookup_keys_ordered.append((style_base, False))
            for lk, is_exact in lookup_keys_ordered:
                lk_key = f"{lk}|{colour_key}"
                dedup_key = (lk_key, entry["colour"], entry["factory"], entry["product_name"], entry["customer_name"])
                if dedup_key in seen_entries:
                    continue
                seen_entries.add(dedup_key)
                if is_exact:
                    result[lk_key].insert(0, entry)  # type: ignore
                else:
                    result[lk_key].append(entry) # type: ignore
    return dict(result)


def _format_date(value: Any, fmt: str) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, datetime):
        return value.strftime(fmt)
    if isinstance(value, date):
        return value.strftime(fmt)
    parsed = _parse_date(value)
    if parsed:
        return parsed.strftime(fmt)
    return ""


def _format_manual_keydate(value: Any) -> str:
    raw = _as_text(value)
    if not raw:
        return ""
    parsed = _parse_date(raw)
    if parsed:
        return f"{parsed.month}/{parsed.day}/{parsed.year}"
    return raw


def _parse_date(value: Any) -> datetime | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())

    # Excel serial number: numeric value in range [1, 2958465] (1900-01-01 to 9999-12-31)
    # Excel epoch is 1899-12-30 (day 0), so serial 1 = 1900-01-01
    if isinstance(value, (int, float)):
        serial = int(value)
        if 1 <= serial <= 2958465:
            from datetime import timedelta
            excel_epoch = datetime(1899, 12, 30)
            return excel_epoch + timedelta(days=serial)
        return None

    raw = str(value).strip()
    if not raw:
        return None

    # Try formats in US-first order (most common in buy files), then international variants.
    # US slash/dash formats are tried before European to avoid ambiguity when day <= 12.
    for c in (
        "%Y-%m-%d",      # ISO: 2026-06-17
        "%Y/%m/%d",      # ISO slash: 2026/06/17
        "%m/%d/%Y",      # US slash: 06/17/2026
        "%m-%d-%Y",      # US dash: 06-17-2026
        "%d-%b-%Y",      # DD-Mon-YYYY: 17-Jun-2026
        "%d-%B-%Y",      # DD-Month-YYYY: 17-June-2026
        "%b %d %Y",      # Mon DD YYYY: Jun 17 2026
        "%B %d %Y",      # Month DD YYYY: June 17 2026
        "%d/%m/%Y",      # European slash: 17/06/2026
    ):
        try:
            return datetime.strptime(raw, c)
        except ValueError:
            continue
    return None


def _transport_method(value: Any) -> str:
    text = _as_text(value)
    key = text.lower()
    return TRANSPORT_MAP.get(key, text or "Sea")


def _format_transport_location(value: Any) -> str:
    raw = _strip_brackets(_as_text(value))
    if not raw:
        return ""
    key = raw.strip().upper()
    return COUNTRY_NAME_MAP.get(key, raw.strip())


def _customer_suffix(raw_customer: str) -> str:
    text = (raw_customer or "").lower()
    if "smu" in text:
        return "SMU"
    if "rto" in text:
        return "RTO"
    if "outlet" in text:
        return "Outlet"
    return ""


def _normalize_vans_po_suffix(raw_customer: str) -> str:
    text = _strip_brackets(_as_text(raw_customer)).strip()
    key = text.lower()
    if not key:
        return ""
    if "south ontario" in key:
        return "South Ontario"
    if "brampton" in key:
        return "Brampton"
    if "sun and sand sports" in key:
        return "Sun and Sand Sports"
    if "vf prague" in key:
        return "VF Prague DC CZ"
    if "vf northern europe" in key:
        return "VF Northern Europe (UK)"
    return re.sub(r"\s+dc$", "", text, flags=re.IGNORECASE).strip()


def _normalize_vans_plant_code(raw_plant: str) -> str:
    plant = _strip_brackets(_as_text(raw_plant)).strip()
    if not plant:
        return ""
    if plant.lower() == "d00028":
        return "VD10"
    return plant.upper()


def _normalize_vans_line_status(po_value: str, status_value: str) -> str:
    return ""


def _normalize_status(raw_status: str, brand: str) -> str:
    status = _strip_brackets(_as_text(raw_status)).strip()
    brand_key = _strip_brackets(_as_text(brand)).strip().lower()
    if brand_key == "vans" and (not status or status.lower() == "converted"):
        return "Confirmed"
    return status or "Confirmed"


def _format_product_range(season: str) -> str:
    normalized = _strip_brackets((season or "").strip())
    # Hunter: Requisition No has format "AW26_UKSOS" — strip everything after underscore
    if "_" in normalized:
        normalized = normalized.split("_")[0].strip()
    # Handle FW26→FH:2026, SS26→SH:2026, AW27→FH:2027, F26→FH:2026
    # A/AW (Autumn/Winter) maps to Fall half (FH), same as F/FW
    match = re.match(r"^(\d{2})FA$", normalized, flags=re.IGNORECASE)
    if match:
        return f"FH:20{match.group(1)}"
    match = re.match(r"^([FSA])(?:W|S)?(\d{2})$", normalized, flags=re.IGNORECASE)
    if match:
        half = "S" if match.group(1).upper() == "S" else "F"
        return f"{half}H:20{match.group(2)}"
    if normalized:
        return normalized
    return "FH:2026"


def _normalize_template(raw_template: str) -> str:
    """Normalise raw doc-type codes from buy files to internal template tokens."""
    normalized = (raw_template or "").strip().upper()
    template_map = {
        "OG": "BULK", "ZNB1": "BULK", "ZMF1": "BULK", "ZDS1": "BULK",
        "SMS": "SMS",
    }
    return template_map.get(normalized, (raw_template or "BULK").strip() or "BULK")


def _resolve_orders_template(brand: str, raw_template: str, brand_config: dict[str, Any] | None = None) -> str:
    """Return the ORDERS Template value for this brand.
    Brand-config value takes priority; then brand-specific map; then _normalize_template()."""
    if brand_config:
        cfg = _as_text(brand_config.get("orders_template"))
        if cfg:
            return cfg
    brand_key = (brand or "").strip().lower()
    if brand_key in BRAND_ORDERS_TEMPLATE_MAP:
        return BRAND_ORDERS_TEMPLATE_MAP[brand_key]
    return _normalize_template(raw_template)


def _resolve_lines_template(brand: str, raw_template: str, brand_config: dict[str, Any] | None = None) -> str:
    """Return the LINES Template value for this brand."""
    if brand_config:
        cfg = _as_text(brand_config.get("lines_template"))
        if cfg:
            return cfg
    brand_key = (brand or "").strip().lower()
    if brand_key in BRAND_LINES_TEMPLATE_MAP:
        return BRAND_LINES_TEMPLATE_MAP[brand_key]
    return _normalize_template(raw_template)


def _build_comments(brand: str, season: str, buy_date: Any, template: str, buy_round: str = "") -> str:
    b = (brand or "OUTPUT").strip() or "OUTPUT"
    s = (season or "NOS").strip() or "NOS"
    parsed = _parse_date(buy_date)
    if buy_round:
        return f"{b} {s} {buy_round} {template}"
    if parsed:
        mon_short = parsed.strftime("%b")
        day = parsed.strftime("%d")
        mon_upper = mon_short.upper()
        suffix = f" {template}" if template else ""
        return f"{b} {s} {mon_short} Buy {day}-{mon_upper}{suffix}"
    return f"{b} {s}"


# Fix #4: parse_int for UDF-buyer_po_number, keep 0 as 0
def parse_int(value):
    try:
        return int(value)
    except Exception:
        return ""

def _to_int_quantity(value: Any) -> int:
    if value in (None, ""):
        return 0
    if isinstance(value, bool):
        return 0
    if isinstance(value, (int, float)):
        return int(value)
    try:
        return int(float(str(value).strip()))
    except ValueError:
        return 0


def _append_row(ws, values: list[Any]) -> None:
    ws.append(["" if v is None else v for v in values])


def _normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip().lower())


def _compact_text(value: str) -> str:
    return re.sub(r"[^a-z0-9]", "", value.lower())


def _strip_brackets(value: str) -> str:
    if not value:
        return value
    # Remove bracket characters but keep the inner text
    cleaned = re.sub(r"\[([^\]]+)\]", r"\1", value)
    cleaned = cleaned.replace("[", "").replace("]", "")
    return re.sub(r"\s+", " ", cleaned).strip()


def _extract_evo_ex_fty(raw_df: Any) -> str:
    """Extract the EVO ex-factory date from the sheet header area."""
    try:
        max_rows = min(10, raw_df.max_row)
        max_cols = min(10, raw_df.max_column)
        for r in range(1, max_rows + 1):
            for c in range(1, max_cols + 1):
                label = _normalize_header(raw_df.cell(row=r, column=c).value)
                if label == "ex factory date":
                    candidate = raw_df.cell(row=r, column=min(c + 1, max_cols)).value
                    formatted = _format_date(candidate, "%m/%d/%Y")
                    return formatted or _as_text(candidate).strip()
    except Exception:
        pass
    return ""


def _detect_evo_pivot_cols(header_series: Any) -> list[int]:
    """Return EVO destination columns from the pivot header row."""
    fixed_headers = {
        "style #", "sku", "upc", "name", "color", "size", "lifecycle", "jpy", "bulk qty",
    }
    pivot_cols: list[int] = []
    for col_idx, value in enumerate(header_series, start=1):
        hdr = _normalize_header(_as_text(value))
        if not hdr or hdr in fixed_headers:
            continue
        pivot_cols.append(col_idx)
    return pivot_cols


def _process_evo_pivot_row(
    raw_df: Any,
    row_idx: int,
    header_row_idx: int,
    destination_row_idx: int,
    or_row_idx: int,
    pivot_cols: list[int],
) -> list[dict[str, Any]]:
    """Explode one EVO source row into per-destination entries."""
    entries: list[dict[str, Any]] = []
    for col_idx in pivot_cols:
        qty = _to_int_quantity(raw_df.cell(row=row_idx, column=col_idx).value)
        if qty <= 0:
            continue
        dest_label = _strip_brackets(_as_text(raw_df.cell(row=destination_row_idx, column=col_idx).value)).strip()
        or_number = _strip_brackets(_as_text(raw_df.cell(row=or_row_idx, column=col_idx).value)).strip()
        if not or_number and not dest_label:
            continue
        if or_number.lower() == "or-xxx":
            continue
        entries.append({
            "qty": qty,
            "destination_label": dest_label,
            "or_number": or_number,
            "column": col_idx,
            "header_row_idx": header_row_idx,
        })
    return entries


def _extract_fox_bracketed_colour(value: str) -> str:
    text = _strip_brackets(_as_text(value)).strip()
    if not text:
        return ""
    matches = re.findall(r"\[([^\]]+)\]", _as_text(value))
    if matches:
        bracketed = matches[-1].strip()
        if bracketed:
            return bracketed
    return text


def _should_silently_ignore_header(header: str) -> bool:
    normalized = _normalize_header(header)
    exact_ignore = {
        "lineitem", "purchaseprice", "sellingprice", "supplierprofile",
        "closeddate", "comments", "currency", "archivedate",
        "productexternalref", "productcustomerref", "purchaseuom",
        "sellinguom", "paymentterm", "defaultdeliverydate",
        "productsupplierext", "keyuser1", "keyuser2", "keyuser3",
        "keyuser4", "keyuser5", "keyuser6", "keyuser7", "keyuser8",
        "department", "customattribute1", "customattribute2",
        "customattribute3", "lineratio", "colourext", "customerext",
        "departmentext", "customattribute1ext", "customattribute2ext",
        "customattribute3ext", "vendor name", "final destination",
        "ship to", "report_date", "seller name", "original po number",
        "po category", "product category", "sku #", "collaboration status",
        "order status", "item status", "original latest date",
        "customs code", "destination name", "upc/ean number", "cost", "sell",
    }
    if normalized in exact_ignore:
        return True
    if normalized.startswith("findfield_") or normalized.startswith("udf-"):
        return True
    if re.match(r"^po[\-_]?\d{4,}$", normalized, re.IGNORECASE):
        return True
    if re.match(r"^\d{5,}$", normalized):
        return True
    return False


def _detect_pivot_format(headers_by_col: dict[int, str], col_map: dict[str, int]) -> dict[str, Any]:
    fixed_cols = sorted({
        col for key, col in col_map.items()
        if col and key not in {"size"}
    })
    max_fixed = max(fixed_cols) if fixed_cols else 0
    pivot_cols: list[tuple[int, str]] = []
    for col, header in sorted(headers_by_col.items()):
        if col <= max_fixed:
            continue
        if not header.strip():
            continue
        normalized = _normalize_header(header)
        if any(_header_matches(normalized, alias) for aliases in HEADER_ALIASES.values() for alias in aliases):
            continue
        if _should_silently_ignore_header(header):
            continue
        pivot_cols.append((col, header))
    has_required_fixed = (("product" in col_map or "product_alt" in col_map) and ("season" in col_map))
    return {
        "is_pivot": bool(pivot_cols and has_required_fixed),
        "pivot_cols": pivot_cols,
        "fixed_cols": fixed_cols,
    }


def _header_matches(header: str, alias: str) -> bool:
    if not header or not alias:
        return False
    if header == alias:
        return True
    h_compact = _compact_text(header)
    a_compact = _compact_text(alias)
    if h_compact == a_compact:
        return True
    alias_has_multiple_tokens = bool(re.search(r"[\s#/\-]", alias))
    if alias_has_multiple_tokens and a_compact and len(a_compact) > 3 and a_compact in h_compact:
        return True
    return False


def _normalize_po(value: Any) -> str:
    if value is None:
        return ""
    # Strip leading/trailing whitespace only — preserve internal spacing
    # (some PO numbers like "F  164860 OG" have intentional double spaces)
    candidate = str(value).strip()
    return candidate


# Numeric factory codes → resolved supplier name
FACTORY_CODE_MAP: dict[str, str] = {
    "508582":   "PT. UWU JUMP INDONESIA",
    "1002436":  "PT. UWU JUMP INDONESIA",
    "8668:puj": "PT. UWU JUMP INDONESIA",  # 511 Tactical
    "mad001":   "PT. UWU JUMP INDONESIA",
    # Obermeyer — different factory
    "hangzhou u-jump arts and crafts": "Hangzhou U-Jump Arts and Crafts",
}


def _resolve_supplier(vendor_code: str, vendor_name: str, brand: str) -> str:
    """Resolve product supplier from available fields, with brand-level fallback."""
    code = (vendor_code or "").strip()
    # Numeric factory codes must be resolved through the map, not passed through raw
    if code in FACTORY_CODE_MAP:
        return FACTORY_CODE_MAP[code]
    if code and len(code) > 2 and not code.isdigit():
        return code
    if vendor_name and len(vendor_name) > 2:
        return vendor_name.strip()
    brand_key = (brand or "").strip().lower()
    return BRAND_SUPPLIER_MAP.get(brand_key, "MISSING_SUPPLIER")


def _resolve_customer(customer_raw: str, brand: str, fallback: str) -> str:
    """Resolve output customer name from available fields."""
    customer_raw = _strip_brackets(customer_raw)
    brand = _strip_brackets(brand)
    if customer_raw:
        brand_key = customer_raw.strip().lower()
        mapped = BRAND_CUSTOMER_MAP.get(brand_key)
        if mapped:
            return mapped
        # If it looks like a code (short, uppercase), try brand map too
        if len(customer_raw) <= 6 and customer_raw.isupper():
            return BRAND_CUSTOMER_MAP.get(brand_key, customer_raw)
        return customer_raw.strip()
    if brand:
        brand_key = brand.strip().lower()
        return BRAND_CUSTOMER_MAP.get(brand_key, brand.strip())
    return fallback


def _resolve_customer_subtype(customer_raw: str, brand: str, fallback: str) -> str:
    """Resolve full customer name including In-Line / SMU / RTO subtype.

    Priority:
    1. Direct match in TNF_CUSTOMER_SUBTYPE_MAP (handles raw values already
       containing the subtype signal, e.g. 'The North Face RTO').
    2. BRAND_CUSTOMER_MAP default (In-Line) for the given brand.
    3. customer_raw as-is if non-empty.
    4. fallback.
    """
    customer_raw = _strip_brackets(customer_raw)
    brand = _strip_brackets(brand)
    if customer_raw:
        key = customer_raw.strip().lower()
        # Try explicit subtype map first
        if key in TNF_CUSTOMER_SUBTYPE_MAP:
            return TNF_CUSTOMER_SUBTYPE_MAP[key]
        # Try brand customer map
        brand_key = key
        if brand_key in BRAND_CUSTOMER_MAP:
            return BRAND_CUSTOMER_MAP[brand_key]
        # Short uppercase code → try brand map
        if len(customer_raw) <= 6 and customer_raw.isupper():
            brand_key2 = customer_raw.strip().lower()
            return BRAND_CUSTOMER_MAP.get(brand_key2, customer_raw)
        return customer_raw.strip()
    if brand:
        brand_key = brand.strip().lower()
        return BRAND_CUSTOMER_MAP.get(brand_key, brand.strip())
    return fallback


def _resolve_keyusers(brand: str) -> dict[str, str]:
    """Return KeyUser1-8 dict for the given brand."""
    brand_key = (brand or "").strip().lower()
    return dict(BRAND_KEYUSER_MAP.get(brand_key, _DEFAULT_KEYUSERS))


# ─────────────────────────────────────────────────────────────────────────────
# Output slice validation  (NEW)
# Validates the three generated workbooks against the reference format rules.
# Returns a list of (level, message) tuples — level is "ERROR" or "WARNING".
# ─────────────────────────────────────────────────────────────────────────────

def _validate_output_slices(
    orders_ws,
    lines_ws,
    sizes_ws,
    validate_sizes: bool = True,
) -> list[tuple[str, str]]:
    """
    Post-generation validation of the three output slices.

    Checks enforced
    ---------------
    1.  Row counts  : LINES rows == SIZES rows (data rows, excluding header)
    2.  Column counts: each sheet has the expected number of columns
    3.  PO universe : ORDERS, LINES, SIZES all share the exact same PO set
    4.  ORDERS      : one row per unique PO (no duplicates)
    5.  Qty total   : SIZES Quantity sum > 0
    6.  Transport   : every TransportMethod value in LINES and ORDERS maps to
                      Sea, Air, or Courier (catches unmapped raw values)
    7.  Customer    : no blank Customer values in LINES or ORDERS
    8.  Status      : ORDERS Status column contains only "Confirmed"
    9.  Duplicate PO+LineItem: LINES and SIZES must have no duplicate keys
    10. LINES/SIZES key alignment: same exact (PO, LineItem) pairs in both sheets
    """
    issues: list[tuple[str, str]] = []

    def col_values(ws, col_idx: int) -> list[str]:
        return [
            _as_text(ws.cell(row=r, column=col_idx).value)
            for r in range(2, ws.max_row + 1)
        ]

    def header_index(ws, name: str) -> int | None:
        for c in range(1, ws.max_column + 1):
            if _as_text(ws.cell(row=1, column=c).value).lower() == name.lower():
                return c
        return None

    # 1. Row counts (optional for ORDER_SIZES)
    lines_data_rows = lines_ws.max_row - 1
    sizes_data_rows = sizes_ws.max_row - 1

    if validate_sizes and lines_data_rows != sizes_data_rows:
        issues.append((
            "ERROR",
            f"Row count mismatch: LINES has {lines_data_rows} rows, "
            f"ORDER_SIZES has {sizes_data_rows} rows."
        ))

    # 2. Column counts
    sheet_checks = [(orders_ws, "ORDERS"), (lines_ws, "LINES")]
    if validate_sizes:
        sheet_checks.append((sizes_ws, "ORDER_SIZES"))
    for ws, sheet_key in sheet_checks:
        expected = EXPECTED_COL_COUNTS[sheet_key]
        actual = ws.max_column
        if actual != expected:
            issues.append((
                "ERROR",
                f"{sheet_key} column count is {actual}, expected {expected}."
            ))

    # 3 & 4. PO universe + ORDERS uniqueness
    orders_po_col = header_index(orders_ws, "PurchaseOrder") or 1
    orders_customer_col = header_index(orders_ws, "Customer")
    lines_po_col  = header_index(lines_ws,  "PurchaseOrder") or 1
    sizes_po_col  = header_index(sizes_ws,  "PurchaseOrder") or 1

    orders_pos = col_values(orders_ws, orders_po_col)
    lines_pos  = col_values(lines_ws,  lines_po_col)
    sizes_pos  = col_values(sizes_ws,  sizes_po_col)

    orders_po_set = set(orders_pos)
    lines_po_set  = set(lines_pos)
    sizes_po_set  = set(sizes_pos)

    if orders_customer_col:
        orders_customers = col_values(orders_ws, orders_customer_col)
        orders_keys = list(zip(orders_pos, orders_customers))
        dupes = [k for k, c in Counter(orders_keys).items() if c > 1]
        if dupes:
            issues.append(("ERROR", f"ORDERS has duplicate PO+Customer keys: {dupes[:5]}"))  # type: ignore
    else:
        dupes = [p for p, c in Counter(orders_pos).items() if c > 1]
        if dupes:
            issues.append(("ERROR", f"ORDERS has duplicate POs: {dupes[:5]}"))  # type: ignore

    if orders_po_set != lines_po_set:
        only_orders = orders_po_set - lines_po_set
        only_lines  = lines_po_set - orders_po_set
        if only_orders:
            issues.append(("ERROR", f"POs in ORDERS but not LINES: {list(only_orders)[:5]}"))  # type: ignore
        if only_lines:
            issues.append(("ERROR", f"POs in LINES but not ORDERS: {list(only_lines)[:5]}"))  # type: ignore

    if validate_sizes and orders_po_set != sizes_po_set:
        only_orders = orders_po_set - sizes_po_set
        only_sizes  = sizes_po_set - orders_po_set
        if only_orders:
            issues.append(("ERROR", f"POs in ORDERS but not ORDER_SIZES: {list(only_orders)[:5]}"))  # type: ignore
        if only_sizes:
            issues.append(("ERROR", f"POs in ORDER_SIZES but not ORDERS: {list(only_sizes)[:5]}"))  # type: ignore

    # 5. Qty total (optional)
    if validate_sizes:
        qty_col = header_index(sizes_ws, "Quantity") or 7
        total_qty = sum(
            _to_int_quantity(sizes_ws.cell(row=r, column=qty_col).value)
            for r in range(2, sizes_ws.max_row + 1)
        )
        if total_qty == 0:
            issues.append(("WARNING", "ORDER_SIZES total Quantity is 0 — all units are zero."))

    # 6. Transport mapping
    for ws, label in [(lines_ws, "LINES"), (orders_ws, "ORDERS")]:
        tm_col = header_index(ws, "TransportMethod")
        if tm_col is None:
            issues.append(("WARNING", f"{label}: TransportMethod column not found."))
            continue
        bad_transport = set()
        for r in range(2, ws.max_row + 1):
            val = _as_text(ws.cell(row=r, column=tm_col).value)
            if val and val not in VALID_TRANSPORT_VALUES:
                bad_transport.add(val)
        if bad_transport:
            issues.append((
                "ERROR",
                f"{label} TransportMethod has unmapped values (must be Sea, Air, or Courier): "
                f"{sorted(bad_transport)}"
            ))

    # 7. Customer not blank
    for ws, label in [(lines_ws, "LINES"), (orders_ws, "ORDERS")]:
        cust_col = header_index(ws, "Customer")
        if cust_col is None:
            issues.append(("WARNING", f"{label}: Customer column not found."))
            continue
        blank_rows = [
            r for r in range(2, ws.max_row + 1)
            if not _as_text(ws.cell(row=r, column=cust_col).value)
        ]
        if blank_rows:
            issues.append((
                "ERROR",
                f"{label} has {len(blank_rows)} blank Customer value(s) "
                f"(first 5 rows: {blank_rows[:5]})."  # type: ignore
            ))

    # 8. Status = Confirmed
    status_col = header_index(orders_ws, "Status")
    if status_col:
        bad_status = set()
        for r in range(2, orders_ws.max_row + 1):
            val = _as_text(orders_ws.cell(row=r, column=status_col).value)
            if val and val != "Confirmed":
                bad_status.add(val)
        if bad_status:
            issues.append((
                "WARNING",
                f"ORDERS Status has unexpected values (expected 'Confirmed'): "
                f"{sorted(bad_status)} — OK if source INFOR view is 'Unconfirmed Only'."
            ))

    # 9. Duplicate PO+LineItem
    li_col_lines = header_index(lines_ws, "LineItem") or 2
    li_col_sizes = header_index(sizes_ws, "LineItem") or 2

    lines_keys = [
        (col_values(lines_ws, lines_po_col)[i], col_values(lines_ws, li_col_lines)[i])
        for i in range(lines_data_rows)
    ]
    sizes_keys = [
        (col_values(sizes_ws, sizes_po_col)[i], col_values(sizes_ws, li_col_sizes)[i])
        for i in range(sizes_data_rows)
    ]

    lines_key_set = set(lines_keys)
    sizes_key_set = set(sizes_keys)

    if len(lines_keys) != len(lines_key_set):
        issues.append(("ERROR", "LINES has duplicate (PurchaseOrder, LineItem) keys."))
    if validate_sizes and len(sizes_keys) != len(sizes_key_set):
        issues.append(("ERROR", "ORDER_SIZES has duplicate (PurchaseOrder, LineItem) keys."))

    # 10. LINES/SIZES key alignment (optional)
    if validate_sizes:
        only_lines_keys = lines_key_set - sizes_key_set
        only_sizes_keys = sizes_key_set - lines_key_set
        if only_lines_keys:
            issues.append((
                "ERROR",
                f"(PO, LineItem) keys in LINES but not ORDER_SIZES: "
                f"{sorted(only_lines_keys)[:5]}"  # type: ignore
            ))
        if only_sizes_keys:
            issues.append((
                "ERROR",
                f"(PO, LineItem) keys in ORDER_SIZES but not LINES: "
                f"{sorted(only_sizes_keys)[:5]}"  # type: ignore
            ))

    return issues


# ─────────────────────────────────────────────────────────────────────────────
# Layout detection  –  scans up to row 80 for the best header row
# ─────────────────────────────────────────────────────────────────────────────

def _detect_layout(
    ws,
    required_keys: set[str] | None = None,
    source_name: str | None = None,
) -> tuple[int, dict[str, int], str, int, int, set[str], dict[str, Any]]:
    best_row = 14
    best_map: dict[str, int] = {}
    best_score = -1
    best_nonempty_po_rows = 0
    best_pivot_info: dict[str, Any] = {"is_pivot": False, "pivot_cols": [], "fixed_cols": []}

    for row_idx in range(1, min(81, ws.max_row + 1)):
        headers_by_col: dict[int, str] = {}
        for col in range(1, ws.max_column + 1):
            normalized = _normalize_header(ws.cell(row=row_idx, column=col).value)
            if normalized:
                headers_by_col[col] = normalized

        if not headers_by_col:
            continue

        # Headers that must never be mapped to product_alt (e.g. "Unique" = JDE+Color concat)
        PRODUCT_ALT_EXCLUDED_HEADERS = {"unique"}
        # Headers that must never be mapped to product (line sequence numbers, not style codes)
        PRODUCT_EXCLUDED_HEADERS = {"item number of purchasing document", "sales order item"}
        # Headers that must never be mapped to orig_ex_fac (often blank, wrong date source)
        ORIG_EX_FAC_EXCLUDED_HEADERS = {"so requested delivery date"}
        # Headers that must never be mapped to vendor_name (distributor name, not actual supplier)
        # Fox Racing col 3 = "Vendor Name" (Madison 88 distributor); col 5 = "Goods Supplier Name" (actual supplier)
        VENDOR_NAME_EXCLUDED_HEADERS = {"vendor name"}
        # Headers that must never be mapped to transport_location (internal DC codes, not country names)
        TRANSPORT_LOCATION_EXCLUDED_HEADERS = {"final destination"}  # Burton internal DC codes
        # Headers that must never be mapped to anything (full address strings, not fields)
        GLOBALLY_IGNORED_HEADERS = {"ship to"}  # prAna: full warehouse address

        col_map: dict[str, int] = {}
        for key, aliases in HEADER_ALIASES.items():
            for col, hdr in headers_by_col.items():
                if key == "product_alt" and hdr in PRODUCT_ALT_EXCLUDED_HEADERS:
                    continue
                if key == "product" and hdr in PRODUCT_EXCLUDED_HEADERS:
                    continue
                if key == "orig_ex_fac" and hdr in ORIG_EX_FAC_EXCLUDED_HEADERS:
                    continue
                if key == "vendor_name" and hdr in VENDOR_NAME_EXCLUDED_HEADERS:
                    continue
                if key == "transport_location" and hdr in TRANSPORT_LOCATION_EXCLUDED_HEADERS:
                    continue
                if hdr in GLOBALLY_IGNORED_HEADERS:
                    continue
                if any(_header_matches(hdr, alias) for alias in aliases):
                    col_map[key] = col
                    break

        dc_plant_cols = [col for col, hdr in headers_by_col.items() if _header_matches(hdr, "dc plant")]
        if dc_plant_cols:
            col_map.setdefault("plant", dc_plant_cols[0])
            if len(dc_plant_cols) > 1:
                col_map.setdefault("plant_name", dc_plant_cols[1])

        # Detect pre-computed NG PO in last column (ON AG, Cotopaxi, 66North, Hunter, Obermeyer, etc.)
        # Pattern: last column header is "PO002924" or similar — the NG PO number itself.
        # Values in that column are the full pre-built PO strings (e.g. "PO002924-SWITZERLAND-ZRH-MKT").
        # Override the "po" mapping to point at this column so the pre-built value is used directly.
        if headers_by_col:
            last_col = max(headers_by_col.keys())
            last_hdr = headers_by_col[last_col]
            if re.match(r'^po\d{4,}$', last_hdr, re.IGNORECASE):
                col_map["po"] = last_col

        pivot_info = _detect_pivot_format(headers_by_col, col_map)
        # Ensure required_keys is a set for iteration
        keys_to_check = required_keys if required_keys is not None else {"po", "qty", "product"}
        # Require at minimum: required keys with special handling for product field
        has_required = True
        for key in cast(set[str], keys_to_check):
            if key == "product":
                if not (("product" in col_map) or ("product_alt" in col_map)):
                    has_required = False
                    break
            elif key in {"po", "qty"} and pivot_info.get("is_pivot"):
                continue
            elif key not in col_map:
                has_required = False
                break
        if not has_required:
            continue

        source_name_lower = (source_name or "").lower()
        if source_name_lower and "511 tactical" in source_name_lower and "quantity" not in col_map:
            for col, hdr in headers_by_col.items():
                if hdr == "total":
                    col_map["quantity"] = col
                    break
        if source_name_lower and "fox racing" in source_name_lower and "quantity" not in col_map:
            for col, hdr in headers_by_col.items():
                if hdr == "sum of order qty":
                    col_map["quantity"] = col
                    break

        score = len(col_map)

        # Probe how many rows below actually have PO data (if PO exists)
        if "po" in col_map:
            po_col = col_map["po"]
            probe_end = min(ws.max_row, row_idx + 300)
            nonempty_po_rows = sum(
                1 for r in range(row_idx + 1, probe_end + 1)
                if _normalize_po(ws.cell(row=r, column=po_col).value)
            )
        else:
            nonempty_po_rows = 0

        if (score > best_score) or (score == best_score and nonempty_po_rows > best_nonempty_po_rows):
            best_score = score
            best_nonempty_po_rows = nonempty_po_rows
            best_row = row_idx
            best_map = col_map
            best_pivot_info = pivot_info

    if best_score >= 3:
        mode = f"auto-detected headers on row {best_row}"
        # Only merge DEFAULT_COL_MAP keys that were NOT detected from actual headers,
        # AND only when the default column index has a sensible header in this file.
        # This prevents COL-specific fixed positions from polluting other brand files.
        merged_map = dict(best_map)  # start from detected keys only
        return best_row + 1, merged_map, mode, best_score, best_nonempty_po_rows, set(best_map.keys()), best_pivot_info # type: ignore

    # Legacy fallback: headers row 14, data row 15
    return 15, DEFAULT_COL_MAP.copy(), "fallback fixed-column layout", 0, 0, set(DEFAULT_COL_MAP.keys()), {"is_pivot": False, "pivot_cols": [], "fixed_cols": []}


def _pick_source_sheet(
    wb,
    requested_sheet: str | None,
    required_keys: set[str] | None = None,
    source_name: str | None = None,
):
    if requested_sheet:
        if requested_sheet not in wb.sheetnames:
            available = ", ".join(wb.sheetnames)
            raise ValueError(f"Sheet '{requested_sheet}' not found. Available: {available}")
        ws = wb[requested_sheet]
        return (ws,) + _detect_layout(ws, required_keys, source_name)

    best_ws = wb.active
    best_data_start = 15
    best_col_map = DEFAULT_COL_MAP.copy()
    best_layout_mode = "fallback fixed-column layout"
    best_score = -1
    best_nonempty = -1
    best_detected: set[str] = set()
    best_pivot_info: dict[str, Any] = {"is_pivot": False, "pivot_cols": [], "fixed_cols": []}

    for ws in wb.worksheets:
        data_start, col_map, layout_mode, score, nonempty, detected, pivot_info = _detect_layout(ws, required_keys, source_name)
        if (score > best_score) or (score == best_score and nonempty > best_nonempty):
            best_ws, best_data_start, best_col_map = ws, data_start, col_map
            best_layout_mode, best_score, best_nonempty, best_detected = layout_mode, score, nonempty, detected
            best_pivot_info = pivot_info

    return best_ws, best_data_start, best_col_map, best_layout_mode, best_score, best_nonempty, best_detected, best_pivot_info


# ─────────────────────────────────────────────────────────────────────────────
# Main processing function
# ─────────────────────────────────────────────────────────────────────────────

def generate_templates(
    input_path: Path,
    output_dir: Path,
    sheet_name: str | None = None,
    customer_fallback: str = DEFAULT_CUSTOMER,
    manual_customer: str | None = None,
    manual_brand: str | None = None,
    strict: bool = False,
    manual_po: str | None = None,
    manual_destination: str | None = None,
    manual_product_range: str | None = None,
    manual_template: str | None = None,
    manual_keydate: str | None = None,
    product_sheet_path: Path | None = None,
    validate_sizes: bool = True,
) -> None:
    wb = load_workbook(input_path, data_only=True)
    source_name = input_path.name.lower()
    manual_po_norm = _normalize_po(manual_po) if manual_po else ""
    manual_product_range = (manual_product_range or "").strip() or ""
    manual_destination = (manual_destination or "").strip() or ""
    manual_template = (manual_template or "").strip() or ""
    manual_keydate = (manual_keydate or "").strip() or ""
    # manual_customer overrides customer for every row (e.g. 511 Tactical has no customer column)
    if manual_customer:
        customer_fallback = manual_customer.strip()
    elif "vuori" in source_name:
        customer_fallback = "Vuori"
    default_qty_if_missing = bool(manual_po_norm)

    product_sheet_map: dict[str, list[dict[str, Any]]] = {}
    if product_sheet_path and product_sheet_path.exists():
        product_wb = load_workbook(product_sheet_path, data_only=True)
        product_sheet_map = _extract_product_sheet_map_from_wb(product_wb)
    else:
        product_sheet_map = _extract_product_sheet_map_from_wb(wb)

    required_keys = {"product"}
    if not manual_po_norm:
        required_keys.add("po")
    if not default_qty_if_missing:
        required_keys.add("qty")

    src, data_start_row, col_map, layout_mode, layout_score, probed_po_rows, detected_keys, pivot_info = \
        _pick_source_sheet(wb, sheet_name, required_keys, source_name)
    evo_ex_fty = _extract_evo_ex_fty(src) if "evo" in source_name else ""

    orders_wb, lines_wb, sizes_wb = Workbook(), Workbook(), Workbook()
    orders_ws = orders_wb.active
    lines_ws  = lines_wb.active
    sizes_ws  = sizes_wb.active
    orders_ws.title, lines_ws.title, sizes_ws.title = "ORDERS", "LINES", "ORDER_SIZES"

    _append_row(orders_ws, ORDERS_HEADERS)
    _append_row(lines_ws,  LINES_HEADERS)
    _append_row(sizes_ws,  SIZES_HEADERS)

    seen_orders:         set[tuple[str, str]]  = set()
    line_item_counter:   dict[str, int]        = defaultdict(int)
    po_to_lines:         dict[str, list[int]]  = defaultdict(list)
    po_to_nonzero_lines: dict[str, list[int]]  = defaultdict(list)
    po_to_sizes:         dict[str, list[int]]  = defaultdict(list)

    total_buy_rows = total_lines_rows = total_sizes_rows = 0
    skipped_no_colour = 0
    skipped_empty_po = 0
    skipped_empty_po_samples: list[tuple] = []
    validation_warnings: list[str] = []
    validation_errors:   list[str] = []
    qty_default_warned = False

    def add_warning(msg: str) -> None:
        if len(validation_warnings) < 200:
            validation_warnings.append(msg)

    def add_error(msg: str) -> None:
        if len(validation_errors) < 200:
            validation_errors.append(msg)

    # Column-level checks
    if "product" not in detected_keys and "product_alt" not in detected_keys:
        msg = "Missing required column mapping for 'product' (header not detected)."
        (add_error if strict else add_warning)(msg)
    if not manual_po_norm and "po" not in detected_keys and not pivot_info.get("is_pivot"):
        msg = "Missing required column mapping for 'po' (header not detected)."
        (add_error if strict else add_warning)(msg)
    if not manual_product_range and "season" not in detected_keys:
        msg = "Missing required column mapping for 'season' (header not detected)."
        (add_error if strict else add_warning)(msg)
    if not default_qty_if_missing and "qty" not in detected_keys and not pivot_info.get("is_pivot"):
        msg = "Missing required column mapping for 'qty' (header not detected)."
        (add_error if strict else add_warning)(msg)

    if "brand" not in detected_keys:
        add_warning("Missing column mapping for 'brand' (comments will use fallback brand value).")

    def _cell(row_idx: int, key: str) -> Any:
        col_idx = col_map.get(key) # type: ignore
        if not col_idx:
            return None
        return src.cell(row=row_idx, column=col_idx).value # type: ignore

    warned_empty_plant_destination: set[str] = set()

    assert src.max_row is not None
    for row_idx in range(data_start_row, src.max_row + 1): # type: ignore
        pivot_expansions = [{"pivot_col": None, "pivot_header": "", "pivot_qty": None}]
        if pivot_info.get("is_pivot"):
            pivot_expansions = []
            for pivot_col, pivot_header in pivot_info.get("pivot_cols", []):  # type: ignore
                pivot_qty = _to_int_quantity(src.cell(row=row_idx, column=pivot_col).value) # type: ignore
                if pivot_qty > 0:
                    pivot_expansions.append({  # type: ignore
                        "pivot_col": pivot_col,
                        "pivot_header": pivot_header,
                        "pivot_qty": pivot_qty,
                    })
            if not pivot_expansions:
                continue

        for pivot_expansion in pivot_expansions:
            def _row_cell(key: str) -> Any:
                if pivot_expansion["pivot_col"] is not None:
                    if key == "qty":
                        return pivot_expansion["pivot_qty"]
                    if key == "po" and not manual_po_norm:
                        return pivot_expansion["pivot_header"]
                    if key == "transport_location" and not manual_destination and "transport_location" not in col_map:
                        return pivot_expansion["pivot_header"]
                return _cell(row_idx, key)

            raw_po_val = _row_cell("po")
            po = manual_po_norm or _normalize_po(raw_po_val)
            if not po:
                skipped_empty_po += 1 # type: ignore
                if len(skipped_empty_po_samples) < 10:
                    skipped_empty_po_samples.append((
                        row_idx,
                        _as_text(src.cell(row=row_idx, column=1).value), # type: ignore
                        _as_text(src.cell(row=row_idx, column=2).value),
                        _as_text(src.cell(row=row_idx, column=3).value),
                        _as_text(raw_po_val),
                    ))
                continue

            # ── Core fields ──────────────────────────────────────────────────────
            product = _as_text(_row_cell("product")) or _as_text(_row_cell("product_alt"))
        if not product and not product_sheet_map:
            add_warning(f"Row {row_idx} PO {po}: product is empty; row skipped.")
            continue
        if not product and product_sheet_map:
            product = ""
        product_external_ref = _as_text(_cell(row_idx, "product_external_ref"))
        product_customer_ref = _as_text(_cell(row_idx, "product_customer_ref"))
        brand_value  = _as_text(_cell(row_idx, "brand")) or (manual_brand or "")
        if not brand_value and "vuori" in source_name:
            brand_value = "vuori"
        if not brand_value and "evo" in source_name:
            brand_value = "evo"
        colour  = _as_text(_cell(row_idx, "colour"))
        colour_display = _as_text(_cell(row_idx, "colour_display"))  # Longtext: human-readable name
        if (brand_value or "").strip().lower() == "vuori" and colour_display:
            colour = colour_display
        if (brand_value or "").strip().lower() == "fox racing":
            fox_colour = _extract_fox_bracketed_colour(colour)
            if fox_colour:
                colour = fox_colour
        if (brand_value or "").strip().lower() == "evo" and evo_ex_fty:
            orig_ex_fac = evo_ex_fty
        qty_cell = _cell(row_idx, "qty")
        if qty_cell is None and default_qty_if_missing:
            qty = 1
            if not qty_default_warned:
                add_warning("Quantity column missing; defaulting Quantity=1 for all rows.")
                qty_default_warned = True
        else:
            qty = _to_int_quantity(qty_cell)

        buyer_po_number_raw = _cell(row_idx, "buyer_po_number")
        vans_confirmed_vendor_crd = _cell(row_idx, "vans_confirmed_vendor_crd")
        brand_requested_crd = _cell(row_idx, "brand_requested_crd")
        is_vans_row = _as_text(_cell(row_idx, "brand")).strip().lower() == "vans" or BRAND_CUSTOMER_MAP.get(_as_text(_cell(row_idx, "customer")).strip().lower()) == "Vans"
        orig_ex_fac  = (vans_confirmed_vendor_crd or brand_requested_crd or _cell(row_idx, "orig_ex_fac") or _cell(row_idx, "confirmed_ex_fac")) if is_vans_row else (_cell(row_idx, "orig_ex_fac") or _cell(row_idx, "confirmed_ex_fac"))
        buy_date     = _cell(row_idx, "buy_date")
        trans_cond   = _cell(row_idx, "trans_cond")
        cancel_date_raw = _cell(row_idx, "cancel_date")
        # Fix #1: Use actual season/range value, raise error if missing
        season_raw = _as_text(_cell(row_idx, "season"))
        season_value = manual_product_range or season_raw
        if not season_value:
            add_warning(f"Row {row_idx} PO {po}: season/range is empty; row skipped.")
            continue
        template_raw = _as_text(_cell(row_idx, "template"))
        customer_raw = _as_text(_cell(row_idx, "customer"))
        if not brand_value and customer_raw:
            customer_key = customer_raw.strip().lower()
            if BRAND_CUSTOMER_MAP.get(customer_key) == "Vans" or "vans" in customer_key:
                brand_value = "vans"
        size_raw     = _as_text(_cell(row_idx, "size"))
        vendor_code  = _as_text(_cell(row_idx, "vendor_code"))
        vendor_name  = _as_text(_cell(row_idx, "vendor_name"))
        buy_round    = _as_text(_cell(row_idx, "submit_buy"))
        status_raw   = _as_text(_cell(row_idx, "status"))

        if not colour or colour.strip().lower() == "not set":
            skipped_no_colour += 1
            add_warning(f"Row {row_idx} PO {po}: colour is empty or 'Not Set'; line/size skipped.")
            continue

        plm_entry = None
        colour_out = colour
        plm_missing = False
        if product_sheet_map:
            colour_key = _normalize_colour_key(colour)
            jde_style_raw = _as_text(_cell(row_idx, "product_alt"))
            jde_style = _normalize_style_key(jde_style_raw)
            # Fall back to product field when product_alt (JDE Style) is absent
            # e.g. ON AG INFOR uses Buyer Item # as product, no separate JDE Style column
            if not jde_style:
                product_raw = _as_text(_cell(row_idx, "product"))
                jde_style = _normalize_style_key(product_raw)
            if not jde_style:
                add_warning(f"Row {row_idx} PO {po}: JDE Style missing; PLM fields left blank.")
                plm_missing = True
            lookup_key = f"{jde_style}|{colour_key}" if jde_style else ""
            matches = product_sheet_map.get(lookup_key, []) if lookup_key else [] # type: ignore
            # If multiple matches, use the first (exact buyer_style_number matches are inserted first)
            if len(matches) > 1:
                matches = [matches[0]]
            if len(matches) == 0 and jde_style:
                style_colour_code = _extract_style_colour_code(jde_style)
                fallback_key = f"{jde_style}|{style_colour_code}" if style_colour_code else ""
                matches = product_sheet_map.get(fallback_key, []) if fallback_key else [] # type: ignore
                if len(matches) > 1:
                    matches = [matches[0]]
            # Prefix match: PLM Buyer Style Number may be a prefix of the buy file style key
            # e.g. PLM has "2UF1067", buy file has "2UF10674959" — try all PLM keys where
            # the style portion is a prefix of jde_style
            # Also handles colour key mismatch via contains check (e.g. "espresso" in "on 003 espresso")
            if len(matches) == 0 and jde_style and colour_key:
                for plm_key, plm_entries in product_sheet_map.items():
                    if "|" not in plm_key:
                        continue
                    plm_style, plm_colour = plm_key.rsplit("|", 1)
                    if not plm_style:
                        continue
                    style_ok = (jde_style.upper().startswith(plm_style.upper()) or
                                plm_style.upper().startswith(jde_style.upper()))
                    if not style_ok:
                        continue
                    colour_ok = (plm_colour == colour_key or
                                 plm_colour in colour_key or colour_key in plm_colour)
                    if not colour_ok and plm_entries:
                        # Raw colour word match (e.g. "espresso" in "on 003 espresso")
                        plm_colour_raw = _as_text(plm_entries[0].get("colour")).lower()
                        colour_ok = (colour.lower() in plm_colour_raw or
                                     plm_colour_raw in colour.lower())
                    if colour_ok:
                        matches = plm_entries[:1]  # type: ignore
                        break
            if len(matches) == 0 and not plm_missing:
                add_warning(f"Row {row_idx} PO {po}: JDE {jde_style} color {colour} not found in PLM sheet; PLM fields left blank.")
                plm_missing = True
            if not plm_missing and len(matches) == 1:
                plm_entry = matches[0]
            if plm_entry and _as_text(plm_entry.get("colour")).strip().lower() == "not set":
                skipped_no_colour += 1
                add_warning(f"Row {row_idx} PO {po}: PLM Color Name is 'Not Set'; line/size skipped.")
                continue

            bv_lower = brand_value.strip().lower() if brand_value else ""
            if plm_entry and bv_lower != "vans" and _as_text(plm_entry.get("product_name")):
                product = _as_text(plm_entry.get("product_name"))
            if plm_entry and _as_text(plm_entry.get("factory")):
                vendor_code = _as_text(plm_entry.get("factory"))
            if plm_entry and bv_lower != "vans" and _as_text(plm_entry.get("colour")): # type: ignore
                colour_out = _as_text(plm_entry.get("colour"))
        # colour_out is PLM Color Name if found, else raw Material value — Longtext (colour_display) is NOT used as colour output

        # Build PO suffix: ManualPO-PlantCode-PlantName (M88) or ManualPO-Plant-Dest (other files)
        # Skip suffix building if PO came from the pre-computed last column (already fully formed,
        # e.g. ON AG "PO002924-SWITZERLAND-ZRH-MKT", Cotopaxi "PO002864", Hunter "PO002933-UKSOS")
        po_col_idx = col_map.get("po")
        po_is_precomputed = (po_col_idx is not None and po_col_idx == src.max_column) # type: ignore
        brand_key_for_row = (brand_value or "").strip().lower()
        plant_value_raw = _as_text(_cell(row_idx, "plant"))
        plant_value = _normalize_vans_plant_code(plant_value_raw) if brand_key_for_row == "vans" else plant_value_raw
        plant_name_value = _as_text(_cell(row_idx, "plant_name"))
        vans_po_suffix = _normalize_vans_po_suffix(customer_raw) if brand_key_for_row == "vans" else ""
        # Derive transport location: explicit column → plant name map → plant code map
        plant_derived_country = (
            PLANT_COUNTRY_MAP.get(plant_value_raw.strip().lower(), "")
            or PLANT_COUNTRY_MAP.get(plant_value.strip().lower(), "")
            or PLANT_COUNTRY_MAP.get(plant_name_value.strip().lower(), "")
        )
        rossignol_destination = manual_destination or _as_text(_cell(row_idx, "transport_location")) or plant_derived_country
        dest_country_raw = (
            (manual_destination or plant_derived_country or _as_text(_cell(row_idx, "transport_location"))) if brand_key_for_row == "vans"
            else ("France" if brand_key_for_row == "rossignol" and _as_text(rossignol_destination).strip().upper() == "EU" else rossignol_destination)
        )
        dest_country = COUNTRY_NAME_MAP.get(dest_country_raw.strip().upper(), dest_country_raw) if dest_country_raw else ""
        if not po_is_precomputed:
            if brand_key_for_row == "vans" and (plant_value or vans_po_suffix or plant_name_value):
                po = "-".join([po] + [p for p in [plant_value, vans_po_suffix or plant_name_value] if p])
            elif plant_name_value:
                # M88 format: ManualPO-PlantCode-PlantName
                po = "-".join([po] + [p for p in [plant_value, plant_name_value] if p])
            elif plant_value or dest_country:
                po = "-".join([po] + [p for p in [plant_value, dest_country] if p])
        suffix_source = _as_text(plm_entry.get("customer_name")) if plm_entry else ""
        suffix = _customer_suffix(suffix_source or customer_raw or brand_value)
        if suffix and not po.lower().endswith(f" {suffix.lower()}"):
            po = f"{po} {suffix}"

        total_buy_rows += 1

        # ── Derived values ───────────────────────────────────────────────────
        trans_method       = _transport_method(trans_cond)
        if not _as_text(orig_ex_fac):
            add_warning(f"Row {row_idx} PO {po}: exFtyDate is empty; delivery/cancel dates left blank.")
        key_date_obj       = _parse_date(buy_date)
        key_date_lines     = _format_date(buy_date, "%m/%d/%Y")
        delivery_date      = _format_date(orig_ex_fac, "%m/%d/%Y")
        cancel_date        = _format_date(cancel_date_raw or orig_ex_fac, "%m/%d/%Y")

        brand_lookup       = brand_value or customer_raw
        # Vendor name fallback for NF0 rows where brand and customer columns are both empty
        if not brand_lookup:
            _vname_lower = (vendor_name or "").strip().lower()
            if "uwu jump" in _vname_lower or "madison 88" in _vname_lower:
                brand_lookup = "tnf"
        if not brand_lookup and re.match(r"^RL[A-Z0-9]", _as_text(_cell(row_idx, "product")).strip(), flags=re.IGNORECASE):
            brand_lookup = "rossignol"
        # Customer name → brand inference for files with no explicit brand column
        if not brand_value and customer_raw:
            _cust_lower = customer_raw.strip().lower()
            if "haglofs" in _cust_lower or "häglofs" in _cust_lower:
                brand_lookup = "haglofs"
            elif "fox racing" in _cust_lower or "fox" in _cust_lower:
                brand_lookup = "fox racing"
            elif "vuori" in _cust_lower:
                brand_lookup = "vuori"
            elif "511 tactical" in _cust_lower or "511tactical" in _cust_lower:
                brand_lookup = "511 tactical"
            elif "obermeyer" in _cust_lower:
                brand_lookup = "obermeyer"
            elif "on ag" in _cust_lower or "on running" in _cust_lower:
                brand_lookup = "on ag"
            elif "66 degrees north" in _cust_lower or "66north" in _cust_lower:
                brand_lookup = "66 degrees north"
            elif "peak performance" in _cust_lower:
                brand_lookup = "peak performance"
            elif "prana" in _cust_lower:
                brand_lookup = "prana"
            elif "burton" in _cust_lower:
                brand_lookup = "burton"
            elif "cotopaxi" in _cust_lower:
                brand_lookup = "cotopaxi"
            elif "hunter" in _cust_lower:
                brand_lookup = "hunter"
        brand_config       = _get_brand_config(brand_lookup)
        # When customer_raw is empty (e.g. NF0 rows with plm_missing), use brand map instead of raw customer_fallback
        _cust_raw_for_resolve = customer_raw or (BRAND_CUSTOMER_MAP.get((brand_lookup or "").strip().lower(), "") if brand_lookup else "")
        customer_value     = _resolve_customer_subtype(_cust_raw_for_resolve, str(brand_value), customer_fallback)
        supplier_value     = _resolve_supplier(vendor_code, vendor_name, brand_lookup)
        size_value         = "One Size" if (not size_raw or size_raw.strip().lower() in {"os", "ons", "one size"}) else size_raw
        if brand_lookup.strip().lower() in {"vans", "rossignol"}:
            product = _as_text(_cell(row_idx, "product_alt")) or product
        product_range      = _format_product_range(season_value)
        if manual_template:
            orders_template = manual_template
            lines_template = manual_template
        else:
            orders_template = _resolve_orders_template(brand_lookup, template_raw, brand_config)
            lines_template  = _resolve_lines_template(brand_lookup, template_raw, brand_config)
        status_value       = _normalize_status(status_raw, brand_lookup)
        comments_value     = _build_comments(
            brand_lookup, product_range, buy_date, orders_template, buy_round
        )
        keyusers           = _resolve_keyusers(brand_lookup)
        purchase_price     = ""  # purchase price not captured
        key_date_orders    = _format_manual_keydate(manual_keydate) if manual_keydate else (
            _format_date(key_date_obj, "%m/%d/%Y") if key_date_obj else ""
        )
        if plm_entry and _as_text(plm_entry.get("customer_name")):
            customer_value = _as_text(plm_entry.get("customer_name"))
        elif product_sheet_map and not plm_entry:
            # plm_missing: only blank PLM-exclusive fields; keep buy file values
            purchase_price = ""
            # product stays as Material Style from buy file (already set above)
            # colour_out stays as Color from buy file (already set above)
            # vendor_code stays as buy file value (already set above)
            # customer_value: fall back to brand map instead of blanking
            customer_value = _resolve_customer_subtype(_cust_raw_for_resolve, str(brand_value), customer_fallback)

        if brand_config and isinstance(brand_config.get("keyusers"), dict):
            cfg_keyusers = brand_config.get("keyusers") or {}
            keyusers = {
                "KeyUser1": _as_text(cfg_keyusers.get("KeyUser1")) or keyusers.get("KeyUser1", ""),
                "KeyUser2": _as_text(cfg_keyusers.get("KeyUser2")) or keyusers.get("KeyUser2", ""),
                "KeyUser3": _as_text(cfg_keyusers.get("KeyUser3")) or keyusers.get("KeyUser3", ""),
                "KeyUser4": _as_text(cfg_keyusers.get("KeyUser4")) or keyusers.get("KeyUser4", ""),
                "KeyUser5": _as_text(cfg_keyusers.get("KeyUser5")) or keyusers.get("KeyUser5", ""),
                "KeyUser6": _as_text(cfg_keyusers.get("KeyUser6")) or keyusers.get("KeyUser6", ""),
                "KeyUser7": _as_text(cfg_keyusers.get("KeyUser7")) or keyusers.get("KeyUser7", ""),
                "KeyUser8": _as_text(cfg_keyusers.get("KeyUser8")) or keyusers.get("KeyUser8", ""),
            }
        valid_statuses = []
        if brand_config and isinstance(brand_config.get("valid_statuses"), list):
            valid_statuses = [str(s).strip().lower() for s in brand_config.get("valid_statuses") if s]  # type: ignore

        # ── Row-level validation warnings ────────────────────────────────────
        for field_name, field_val, fallback_desc in [
            ("season",   season_value, f"ProductRange '{product_range}'"),
            ("product",  product,      "Product missing"),
            ("template", template_raw, f"Template '{orders_template}'"),
            ("customer", customer_raw, f"Customer '{customer_value}'"),
            ("size",     size_raw,     "Size 'One Size'"),
        ]:
            if not field_val:
                msg = f"Row {row_idx} PO {po}: {field_name} is empty; fallback {fallback_desc} used."
                (add_error if strict else add_warning)(msg)

        if not brand_value:
            add_warning(f"Row {row_idx} PO {po}: brand is empty; comments use fallback brand value.")
        if valid_statuses:
            if status_value and status_value.strip().lower() not in valid_statuses:
                add_warning(
                    f"Row {row_idx} PO {po}: status '{status_value}' not in valid statuses "
                    f"{valid_statuses}."
                )

        # ── ORDERS row (one per unique PO) ───────────────────────────────────
        if brand_lookup.strip().lower() == "evo" and pivot_info.get("is_pivot"):
            evo_header_row_idx = max(1, data_start_row - 1)
            evo_destination_row_idx = max(1, evo_header_row_idx - 2)
            evo_or_row_idx = max(1, evo_header_row_idx - 1)
            evo_header_series = [cell.value for cell in src[evo_header_row_idx]]  # type: ignore
            evo_pivot_cols = _detect_evo_pivot_cols(evo_header_series)
            if not evo_pivot_cols and pivot_info.get("pivot_cols"):
                evo_pivot_cols = [int(col) for col, _ in pivot_info.get("pivot_cols", [])]  # type: ignore
            evo_entries = _process_evo_pivot_row(
                src,
                row_idx,
                evo_header_row_idx,
                evo_destination_row_idx,
                evo_or_row_idx,
                evo_pivot_cols,
            )
            if not evo_entries:
                total_buy_rows -= 1
                continue
            if len(evo_entries) > 1:
                total_buy_rows += len(evo_entries) - 1

            evo_order_customer_key = customer_value or customer_fallback
            for entry in evo_entries:
                evo_po = _normalize_po(entry.get("or_number")) or po
                if not evo_po:
                    continue
                evo_destination = _format_transport_location(
                    entry.get("destination_label")
                    or manual_destination
                    or plant_derived_country
                    or _cell(row_idx, "transport_location")
                    or ""
                )
                evo_order_key = (evo_po, f"{evo_order_customer_key}||{evo_destination}")
                if evo_order_key not in seen_orders:
                    seen_orders.add(evo_order_key)
                    _append_row(orders_ws, [
                        evo_po, supplier_value, status_value, customer_value,
                        trans_method, evo_destination, "", orders_template,
                        key_date_orders,
                        "", "", comments_value, CURRENCY,
                        keyusers["KeyUser1"], keyusers["KeyUser2"], keyusers["KeyUser3"],
                        keyusers["KeyUser4"], keyusers["KeyUser5"], keyusers["KeyUser6"],
                        keyusers["KeyUser7"], keyusers["KeyUser8"],
                        "", "", "", "", "",
                    ])

                line_item_counter[evo_po] += 1 # type: ignore
                line_item = line_item_counter[evo_po]
                _append_row(lines_ws, [
                    evo_po, line_item, product_range, product, customer_value,
                    delivery_date, trans_method, evo_destination, _normalize_vans_line_status(evo_po, status_value), purchase_price, "",
                    lines_template, delivery_date, SUPPLIER_PROFILE,
                    "", "", CURRENCY, "", product_external_ref, product_customer_ref, "", "",
                    evo_po,
                    delivery_date, cancel_date,
                    "", "", "", "", "", "",
                ])
                po_to_lines[evo_po].append(line_item)
                total_lines_rows += 1

                _append_row(sizes_ws, [
                    evo_po, line_item, product_range, product,
                    size_value, size_value, entry["qty"], colour_out,
                    "", "", "", "", "", "", "", "", "", "", "", "", "", "",
                    "", "", "", "", "", "", "",
                ])
                po_to_nonzero_lines[evo_po].append(line_item)
                po_to_sizes[evo_po].append(line_item)
                total_sizes_rows += 1
            continue

        order_customer_key = customer_value or customer_fallback
        order_key = (po, order_customer_key)
        if order_key not in seen_orders:
            seen_orders.add(order_key)
            # Fix #2: Map TransportLocation from source (with plant-derived fallback for M88)
            transport_location = _format_transport_location(
                (manual_destination or plant_derived_country or _cell(row_idx, "transport_location")) if brand_key_for_row == "vans"
                else ("France" if brand_key_for_row == "rossignol" and _as_text(manual_destination or _cell(row_idx, "transport_location") or plant_derived_country).strip().upper() == "EU" else (manual_destination or _cell(row_idx, "transport_location") or plant_derived_country))
            )
            _append_row(orders_ws, [
                po, supplier_value, status_value, customer_value,
                trans_method, transport_location, "", orders_template,
                key_date_orders,
                "", "", comments_value, CURRENCY,
                keyusers["KeyUser1"], keyusers["KeyUser2"], keyusers["KeyUser3"],
                keyusers["KeyUser4"], keyusers["KeyUser5"], keyusers["KeyUser6"],
                keyusers["KeyUser7"], keyusers["KeyUser8"],
                "", "", "", "", "",
            ])

        # Skip LINES / SIZES if Colour is blank
        if not colour:
            skipped_no_colour += 1
            add_warning(f"Row {row_idx} PO {po}: colour is empty; line/size skipped.")
            continue

        # ── LINES row (one per buy file row) ─────────────────────────────────
        line_item_counter[po] += 1 # type: ignore
        line_item = line_item_counter[po]

        # Fix #2: Map TransportLocation from source for LINES (with plant-derived fallback for M88)
        transport_location = _format_transport_location(
            (manual_destination or plant_derived_country or _cell(row_idx, "transport_location")) if brand_key_for_row == "vans"
            else ("France" if brand_key_for_row == "rossignol" and _as_text(manual_destination or _cell(row_idx, "transport_location") or plant_derived_country).strip().upper() == "EU" else (manual_destination or _cell(row_idx, "transport_location") or plant_derived_country))
        )
        # Fix #3: KeyDate per line from DeliveryDate
        key_date_line = delivery_date
        buyer_po_number_out = (
            buyer_po_number_raw if buyer_po_number_raw not in (None, "")
            else _normalize_po(raw_po_val) or ""
        )
        _append_row(lines_ws, [
            po, line_item, product_range, product, customer_value,
            delivery_date, trans_method, transport_location, _normalize_vans_line_status(po, status_value), purchase_price, "",
            lines_template, key_date_line, SUPPLIER_PROFILE,
            "", "", CURRENCY, "", product_external_ref, product_customer_ref, "", "",
            buyer_po_number_out,
            delivery_date, cancel_date,
            "", "", "", "", "", "",
        ])
        po_to_lines[po].append(line_item)
        total_lines_rows += 1

        # ── ORDER_SIZES row (include qty=0) ──────────────────────────────────
        po_to_nonzero_lines[po].append(line_item)
        _append_row(sizes_ws, [
            po, line_item, product_range, product,
            size_value, size_value, qty, colour_out,
            "", "", "", "", "", "", "", "", "", "", "", "", "", "",
            "", "", "", "", "", "", "",
        ])
        po_to_sizes[po].append(line_item)
        total_sizes_rows += 1

    # ── Post-processing integrity checks ─────────────────────────────────────
    unique_order_count = len(seen_orders)
    unique_po_count = len({po for po, _ in seen_orders})
    orders_count    = orders_ws.max_row - 1 # type: ignore

    if orders_count != unique_order_count:
        raise ValueError(
            f"ORDERS row count mismatch. Expected {unique_order_count}, got {orders_count}."
        )

    if total_lines_rows != total_buy_rows - skipped_no_colour: # type: ignore
        raise ValueError(
            f"LINES row count mismatch. Expected {total_buy_rows - skipped_no_colour}, got {total_lines_rows}."
        )

    if total_buy_rows == 0:
        raise ValueError(
            "No usable buy rows detected. Check the sheet/header row, "
            "or pass --sheet with the exact worksheet name."
        )

    for po, line_items in po_to_lines.items():
        expected = list(range(1, len(line_items) + 1))
        if line_items != expected:
            raise ValueError(
                f"LineItem sequence invalid for PO {po}. "
                f"Expected {expected}, got {line_items}."
            )
        size_items          = po_to_sizes.get(po, [])
        expected_size_items = po_to_lines.get(po, [])
        if size_items != expected_size_items:
            raise ValueError(
                f"LineItem mismatch LINES vs ORDER_SIZES for PO {po}. "
                f"Expected {expected_size_items}, got {size_items}."
            )

    if strict and validation_errors:
        _print_summary(
            src, layout_mode, layout_score, probed_po_rows, data_start_row,
            total_buy_rows, unique_po_count, unique_order_count, orders_count,
            total_lines_rows, total_sizes_rows,
            skipped_empty_po, skipped_no_colour, validation_warnings, validation_errors,
            skipped_empty_po_samples,
            validate_sizes=validate_sizes,
        )
        raise ValueError("Strict validation failed due to missing/blank required fields.")

    # ── Output slice validation (NEW) ─────────────────────────────────────────
    slice_issues = _validate_output_slices(
        orders_ws, lines_ws, sizes_ws, validate_sizes=validate_sizes
    )
    slice_errors   = [(lvl, msg) for lvl, msg in slice_issues if lvl == "ERROR"]
    slice_warnings = [(lvl, msg) for lvl, msg in slice_issues if lvl == "WARNING"]

    # ── Write output files ────────────────────────────────────────────────────
    output_dir.mkdir(parents=True, exist_ok=True)
    orders_wb.save(output_dir / "orders.xlsx")
    lines_wb.save(output_dir  / "lines.xlsx")
    sizes_wb.save(output_dir  / "order_sizes.xlsx")

    _print_summary(
        src, layout_mode, layout_score, probed_po_rows, data_start_row,
        total_buy_rows, unique_po_count, unique_order_count, orders_count,
        total_lines_rows, total_sizes_rows,
        skipped_empty_po, skipped_no_colour, validation_warnings, validation_errors,
        skipped_empty_po_samples,
        slice_issues=slice_issues,
        validate_sizes=validate_sizes,
    )

    if slice_errors and strict:
        raise ValueError(
            f"Output slice validation failed with {len(slice_errors)} error(s). "
            "Files written but review the report above."
        )
    print(f"\nGenerated files in: {output_dir.resolve()}")
    for fname in ("orders.xlsx", "lines.xlsx", "order_sizes.xlsx"):
        print(f"  - {fname}")


def _print_summary(
    src, layout_mode, layout_score, probed_po_rows, data_start_row,
    total_buy_rows, unique_po_count, unique_order_count, orders_count,
    total_lines_rows, total_sizes_rows,
    skipped_empty_po, skipped_no_colour, validation_warnings, validation_errors,
    skipped_empty_po_samples,
    validate_sizes: bool = True,
    slice_issues: list[tuple[str, str]] | None = None,
) -> None:
    lines_equals_sizes = total_lines_rows == total_sizes_rows
    slice_issues = slice_issues or []
    slice_errors   = [msg for lvl, msg in slice_issues if lvl == "ERROR"]
    slice_warnings = [msg for lvl, msg in slice_issues if lvl == "WARNING"]
    overall_status = "✓ PASS" if not slice_errors else f"✗ FAIL ({len(slice_errors)} error(s))"

    print("\n── Generation Summary ──────────────────────────────────────────")
    print(f"  Source sheet   : {src.title}")
    print(f"  Layout mode    : {layout_mode}")
    print(f"  Layout score   : {layout_score}")
    print(f"  PO probe rows  : {probed_po_rows}")
    print(f"  Data start row : {data_start_row}")
    print(f"  Buy rows       : {total_buy_rows}")
    print(f"  Unique POs     : {unique_po_count}")
    print(f"  Unique Orders  : {unique_order_count} (PO+Customer)")
    print(f"  ORDERS rows    : {orders_count}")
    print(f"  LINES rows     : {total_lines_rows}")
    if skipped_no_colour:
        print(f"  Skipped (no Colour): {skipped_no_colour}")
    print(f"  ORDER_SIZES    : {total_sizes_rows}")
    if validate_sizes:
        print(f"  LINES==SIZES   : {'YES' if lines_equals_sizes else 'NO (qty=0 rows excluded in ORDER_SIZES)'}")
    else:
        print("  LINES==SIZES   : SKIPPED (sizes validation disabled)")
    print(f"  Skipped (no PO): {skipped_empty_po}")
    print(f"  Warnings       : {len(validation_warnings)}")
    print(f"  Errors         : {len(validation_errors)}")
    if validation_warnings:
        print("  Warning samples:")
        for msg in validation_warnings[:15]:
            print(f"    - {msg}")
    if validation_errors:
        print("  Error samples:")
        for msg in validation_errors[:15]:
            print(f"    - {msg}")
    if skipped_empty_po_samples:
        print("  Empty-PO samples (row, A, B, C, raw PO cell):")
        for row_idx, col_a, col_b, col_c, sample_val in skipped_empty_po_samples:
            print(f"    - {row_idx}: {col_a} | {col_b} | {col_c} | {sample_val}")

    print("\n── Output Slice Validation ─────────────────────────────────────")
    print(f"  Overall        : {overall_status}")
    checks = [
        "Column counts (ORDERS=26, LINES=31)",
        "ORDERS has no duplicate PO+Customer keys",
        "TransportMethod mapped to Sea, Air, or Courier only",
        "No blank Customer values",
        "ORDERS Status = Confirmed",
        "No duplicate PO+LineItem keys",
    ]
    if validate_sizes:
        checks = [
            "Row counts (LINES == SIZES)",
            "Column counts (ORDERS=26, LINES=31, SIZES=29)",
            "PO universe matches across all 3 files",
            "ORDERS has no duplicate PO+Customer keys",
            "ORDER_SIZES total Quantity > 0",
            "TransportMethod mapped to Sea, Air, or Courier only",
            "No blank Customer values",
            "ORDERS Status = Confirmed",
            "No duplicate PO+LineItem keys",
            "LINES/SIZES key alignment",
        ]
    issue_msgs = [msg for _, msg in slice_issues]
    for check in checks:
        hit = any(
            any(keyword in msg for keyword in check.lower().split())
            for msg in issue_msgs
        )
        print(f"  {'✗' if hit else '✓'} {check}")
    if slice_errors:
        print("  Errors:")
        for msg in slice_errors:
            print(f"    ✗ {msg}")
    if slice_warnings:
        print("  Warnings:")
        for msg in slice_warnings:
            print(f"    ⚠ {msg}")
    print("────────────────────────────────────────────────────────────────")


# ─────────────────────────────────────────────────────────────────────────────
# CLI entry point
# ─────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Global NG buy-file processor → ORDERS / LINES / ORDER_SIZES"
    )
    parser.add_argument("--input",      dest="input_file",       required=True)
    parser.add_argument("--output-dir", dest="output_dir",       default=".")
    parser.add_argument("--sheet",      dest="sheet_name",       default=None)
    parser.add_argument("--customer",   dest="customer_fallback", default=DEFAULT_CUSTOMER)
    parser.add_argument("--manual-customer", dest="manual_customer", default=None,
                        help="Force customer name for every row (use for files with no customer column, e.g. 511 Tactical)")
    parser.add_argument("--manual-brand", dest="manual_brand", default=None,
                        help="Fallback brand when file has no brand column (file-level brand takes priority)")
    parser.add_argument("--strict",     dest="strict",           action="store_true")
    parser.add_argument("--po",         dest="manual_po",        default=None)
    parser.add_argument("--destination", dest="manual_destination", default=None)
    parser.add_argument("--product-range", dest="manual_product_range", default=None)
    parser.add_argument("--season",     dest="manual_season",    default=None,
                        help="Alias for --product-range (e.g. FH:2026)")
    parser.add_argument("--template",   dest="manual_template",  default=None)
    parser.add_argument("--keydate",    dest="manual_keydate",   default=None)
    parser.add_argument("--product-sheet", dest="product_sheet", default=None)
    parser.add_argument("--validate-sizes", dest="validate_sizes", action="store_true")
    args = parser.parse_args()

    base           = Path.cwd()
    input_candidate = Path(args.input_file)
    input_path      = input_candidate if input_candidate.is_absolute() else (base / input_candidate)
    if not input_path.exists():
        raise FileNotFoundError(f"Input file not found: {input_path}")

    output_candidate = Path(args.output_dir)
    output_dir       = output_candidate if output_candidate.is_absolute() else (base / output_candidate)
    product_sheet_path = None
    if args.product_sheet:
        candidate = Path(args.product_sheet)
        product_sheet_path = candidate if candidate.is_absolute() else (base / candidate)
        if not product_sheet_path.exists():
            raise FileNotFoundError(f"Product sheet not found: {product_sheet_path}")

    generate_templates(
        input_path        = input_path,
        output_dir        = output_dir,
        sheet_name        = args.sheet_name,
        customer_fallback = args.customer_fallback,
        strict            = args.strict,
        manual_po         = args.manual_po,
        manual_destination = args.manual_destination,
        manual_product_range = args.manual_product_range or args.manual_season,
        manual_template   = args.manual_template,
        manual_keydate    = args.manual_keydate,
        product_sheet_path = product_sheet_path,
        validate_sizes    = args.validate_sizes,
        manual_customer   = args.manual_customer,
        manual_brand      = args.manual_brand,
    )
